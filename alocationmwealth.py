from __future__ import annotations

import json
import math
import re
import unicodedata
from datetime import datetime
from copy import deepcopy
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st

try:
    import yfinance as yf
    HAS_YFINANCE = True
except Exception:
    yf = None
    HAS_YFINANCE = False

import positions as posmod

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    HAS_REPORTLAB = True
except Exception:
    HAS_REPORTLAB = False

PDF_FONT_REGULAR = "Helvetica"
PDF_FONT_BOLD = "Helvetica-Bold"


def register_pdf_fonts() -> tuple[str, str]:
    """Usa fonte Unicode do ambiente para preservar acentos no PDF."""
    global PDF_FONT_REGULAR, PDF_FONT_BOLD
    if not HAS_REPORTLAB:
        return PDF_FONT_REGULAR, PDF_FONT_BOLD
    candidates = [
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf", "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
    ]
    for regular, bold in candidates:
        if Path(regular).exists() and Path(bold).exists():
            try:
                pdfmetrics.registerFont(TTFont("MWRegular", regular))
                pdfmetrics.registerFont(TTFont("MWBold", bold))
                PDF_FONT_REGULAR = "MWRegular"
                PDF_FONT_BOLD = "MWBold"
                break
            except Exception:
                pass
    return PDF_FONT_REGULAR, PDF_FONT_BOLD


st.set_page_config(page_title="M Wealth | Balanceamento", layout="wide", page_icon="📊")

BASE_DIR = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()
POS_DIR = BASE_DIR / "posicoes"
APP_VERSION = "6.2"
DATA_DIR = BASE_DIR / "data"
PUBLISHED_MODELS_PATH = DATA_DIR / "modelos_publicados.json"
MODEL_HISTORY_PATH = DATA_DIR / "historico_modelos.jsonl"


# Estratégia de RV e FiInfra permanece no código, conforme orientação da gestão.
ACOES_SEM_RENDA = ["AXIA3", "EQTL3", "SBSP3", "ITUB3", "BPAC11", "PSSA3", "PRIO3", "VALE3", "WEGE3", "RENT3"]
ACOES_COM_RENDA = ["CPLE3", "EGIE3", "AXIA3", "ITUB3", "VALE3", "ALOS3", "FLRY3", "ABEV3", "PRIO3", "WEGE3"]
FIIS_RECOMENDADOS = ["KNRI11", "XPML11", "HGLG11", "PVBI11", "HGRU11", "KNCR11", "KNIP11", "KNCA11"]
FI_INFRA_TICKERS = ["KNDI11", "CDII11", "IFRI11", "AZIN11", "JURO11", "IFRA11", "KDIF11", "BDIF11", "JMBI11", "CPTI11"]
FI_INFRA_POS_TICKERS = ["KNDI11", "CDII11", "IFRI11", "AZIN11"]
FI_INFRA_INFLACAO_TICKERS = ["JURO11", "IFRA11", "KDIF11", "BDIF11", "JMBI11", "CPTI11"]

# Ativos estratégicos negociados no Brasil que precisam ser tratados pelo código
# antes da regra genérica de ticker terminado em 11, para não caírem como FII.
ATIVOS_ESTRATEGICOS_B3 = {
    # Regras explícitas têm prioridade sobre a heurística genérica de ticker final 11.
    "BPAC11": {"classe": "RV Brasil", "subbucket": "Ações", "estrategia": "Ação / Unit"},
    "NTNS11": {"classe": "RF Brasil", "subbucket": "Inflação - Tesouro", "estrategia": "ETF de renda fixa — Tesouro IPCA 0 a 4 anos"},
    "BITH11": {"classe": "Alternativos", "subbucket": "Bitcoin", "estrategia": "Bitcoin"},
    "GOLD11": {"classe": "Alternativos", "subbucket": "Ouro", "estrategia": "Ouro"},
    "UTLL11": {"classe": "RV Brasil", "subbucket": "Ações", "estrategia": "Utilities"},
    "DIVD11": {"classe": "RV Brasil", "subbucket": "Ações", "estrategia": "Dividendos"},
    "XINA11": {"classe": "Internacional", "subbucket": "Renda Variável Internacional", "estrategia": "Exposição China"},
    "IVVB11": {"classe": "Internacional", "subbucket": "Renda Variável Internacional", "estrategia": "Exposição EUA"},
    "NASD11": {"classe": "Internacional", "subbucket": "Renda Variável Internacional", "estrategia": "Exposição EUA"},
    "SPYI11": {"classe": "Internacional", "subbucket": "Renda Variável Internacional", "estrategia": "Dividendos EUA"},
    "ALUG11": {"classe": "Internacional", "subbucket": "Renda Variável Internacional", "estrategia": "Imobiliário EUA"},
    "USTK11": {"classe": "Internacional", "subbucket": "Renda Variável Internacional", "estrategia": "Tecnologia EUA"},
}

# Mantém ordem operacional das tabelas.
SUBBUCKET_ORDER = [
    "Pós - Imediato", "Pós - 1 a 30 dias", "Pós - 31 a 180 dias", "Pós - 181 a 360 dias", "Pós - 361+ dias",
    "FiInfra Pós", "FiInfra Inflação", "Pré - Bancário", "Pré - Tesouro", "Inflação - Bancário", "Inflação - Tesouro", "Crédito Privado",
    "Ações", "FIIs", "Bitcoin", "Ouro", "Renda Fixa Internacional", "Renda Variável Internacional", "Caixa Internacional",
    "Saldo em Conta", "Proventos a Receber", "Direitos de Subscrição", "Fundos de Investimento / Sem Liquidez Mapeada", "Previdência", "COE / Estruturados", "Outros / Não Classificado",
]

st.markdown(
    """
    <style>
    [data-testid="stSidebar"], [data-testid="collapsedControl"] { display: none !important; }
    .block-container { padding-top: 1.0rem; padding-bottom: 2.0rem; max-width: 1680px; }
    div[data-testid="stMetricValue"] { font-size: 1.2rem; }
    .mw-header { display: flex; align-items: center; gap: 1.2rem; margin-bottom: .35rem; }
    .mw-title { font-size: 2.2rem; line-height: 1.1; font-weight: 800; margin: 0; }
    .mw-version { color: rgba(250,250,250,.45); font-size: .78rem; margin-top: .25rem; }
    .mw-card { border: 1px solid rgba(255,255,255,.10); border-radius: 16px; padding: 16px 18px; background: linear-gradient(180deg, rgba(255,255,255,.055), rgba(255,255,255,.022)); box-shadow: 0 10px 24px rgba(0,0,0,.12); }
    .mw-card-label { color: rgba(250,250,250,.68); font-size: .82rem; font-weight: 650; margin-bottom: .35rem; }
    .mw-card-value { color: #fff; font-size: 1.38rem; font-weight: 800; letter-spacing: -.02em; }
    .mw-muted { color: rgba(250,250,250,.65); font-size: .86rem; }
    .mw-ok { color: #77dd77; font-weight: 700; }
    .mw-warn { color: #ffd166; font-weight: 700; }
    .mw-bad { color: #ff6b6b; font-weight: 700; }
    .mw-line { border-top: 1px solid rgba(255,255,255,.10); margin: .9rem 0 1.1rem 0; }
    .mw-page-intro { border: 1px solid rgba(255,255,255,.10); border-radius: 18px; padding: 20px 22px; margin: 0 0 1rem 0; background: radial-gradient(circle at 85% 15%, rgba(70,104,180,.20), transparent 34%), linear-gradient(135deg, rgba(33,50,84,.75), rgba(14,18,27,.82)); }
    .mw-eyebrow { color: #9fb5e5; font-size: .75rem; font-weight: 800; letter-spacing: .10em; text-transform: uppercase; }
    .mw-page-title { color: #fff; font-size: 1.65rem; font-weight: 850; margin: .20rem 0 .25rem 0; }
    .mw-page-text { color: rgba(250,250,250,.67); max-width: 980px; font-size: .91rem; }
    .mw-section-title { font-size: 1.13rem; font-weight: 800; margin: 1.15rem 0 .20rem 0; }
    .mw-pill { display:inline-block; border:1px solid rgba(255,255,255,.14); border-radius:999px; padding:.25rem .55rem; margin-right:.25rem; color:rgba(255,255,255,.75); font-size:.74rem; }
    div[data-testid="stTabs"] button { font-weight: 750; }
    div[data-testid="stDataFrame"] { border: 1px solid rgba(255,255,255,.08); border-radius: 12px; overflow: hidden; }
    div[data-testid="stExpander"] { border-radius: 12px; }
    </style>
    """,
    unsafe_allow_html=True,
)


# =============================================================================
# Utilitários
# =============================================================================
def find_file(filename: str) -> Path:
    for p in [POS_DIR / filename, BASE_DIR / filename, Path.cwd() / filename, Path(filename)]:
        if p.exists():
            return p
    return POS_DIR / filename

def master_products_path() -> Path:
    """Localiza o cadastro mestre mais recente, inclusive cópias como ``(1)``.

    O Excel/navegador pode salvar o arquivo como ``Cadastro...(1).xlsx``. A
    versão anterior procurava apenas nomes exatos e acabava usando uma base
    antiga ou o fallback ``Nome fundos e prev.xlsx``.
    """
    search_dirs = []
    for directory in [POS_DIR, BASE_DIR, Path.cwd()]:
        try:
            resolved = directory.resolve()
        except Exception:
            resolved = directory
        if resolved not in search_dirs:
            search_dirs.append(resolved)

    patterns = [
        "Cadastro_Mestre_Produtos_M_Wealth*.xlsx",
        "Cadastro Mestre de Produtos M Wealth*.xlsx",
        "Cadastro Mestre de Produtos*.xlsx",
    ]
    candidates: list[Path] = []
    for directory in search_dirs:
        if not directory.exists():
            continue
        for pattern in patterns:
            candidates.extend(
                p for p in directory.glob(pattern)
                if p.is_file() and not p.name.startswith("~$")
            )

    if candidates:
        unique = {str(p.resolve()): p for p in candidates}
        return max(
            unique.values(),
            key=lambda p: (
                p.stat().st_mtime_ns if p.exists() else 0,
                p.stat().st_size if p.exists() else 0,
                p.name,
            ),
        )

    # Compatibilidade com a base legada.
    for filename in ["Nome fundos e prev.xlsx", "Nome fundos e prev (1).xlsx"]:
        p = find_file(filename)
        if p.exists():
            return p
    return find_file("Cadastro_Mestre_Produtos_M_Wealth.xlsx")


def parse_yes_no(value, default: bool = True) -> bool:
    txt = norm(value)
    if txt in {"NAO", "N", "FALSE", "0", "NÃO"}:
        return False
    if txt in {"SIM", "S", "TRUE", "1"}:
        return True
    return bool(default)

def optional_text(value) -> str:
    """Normaliza campos opcionais vindos do Excel, eliminando NaN/None textuais."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    txt = str(value).strip()
    return "" if norm(txt) in {"", "NAN", "NONE", "NULL"} else txt


def canonicalize_master_columns(df: pd.DataFrame, expected: list[str]) -> pd.DataFrame:
    """Corrige cabeçalhos que o Excel renomeou para TICKER2, NOME_ATIVO3 etc.

    Tabelas estruturadas do Excel podem acrescentar números aos cabeçalhos ao
    detectar nomes duplicados. O cadastro continua legível porque comparamos a
    versão normalizada do cabeçalho sem o sufixo numérico.
    """
    out = df.copy()
    expected_by_norm = {norm(col): col for col in expected}
    rename = {}
    for col in out.columns:
        raw = str(col).strip()
        base = re.sub(r"\d+$", "", raw).strip()
        target = expected_by_norm.get(norm(raw)) or expected_by_norm.get(norm(base))
        if target:
            rename[col] = target
    return out.rename(columns=rename)


def logo_path() -> Path | None:
    for p in [BASE_DIR / "Logo-M-Wealth.png", POS_DIR / "Logo-M-Wealth.png", Path.cwd() / "Logo-M-Wealth.png"]:
        if p.exists():
            return p
    return None


def logo_pdf_path() -> Path | None:
    """Logo específica para PDF em fundo claro."""
    for p in [BASE_DIR / "Logo-M-Wealth-Preta.png", POS_DIR / "Logo-M-Wealth-Preta.png", Path.cwd() / "Logo-M-Wealth-Preta.png", BASE_DIR / "Logo-M-Wealth.png"]:
        if p.exists():
            return p
    return None


def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(c))


def norm(x) -> str:
    return strip_accents(str(x or "")).upper().strip()


def ticker_clean(x) -> str:
    return norm(x).replace(" ", "").replace(".", "")


def is_valid_b3_ticker(value) -> bool:
    """Aceita tickers B3 alfanuméricos como B3SA3, BPAC11 e NTNS11."""
    tk = ticker_clean(value)
    return bool(
        5 <= len(tk) <= 7
        and re.fullmatch(r"[A-Z0-9]+", tk)
        and re.search(r"[A-Z]", tk)
        and re.search(r"\d", tk)
    )


def expand_fund_compounds(text: str) -> str:
    """Separa abreviações compostas comuns em nomes de fundos (ex.: FICFIDC, FICFIRF,
    FICFIA, FICFIM), muito usadas por BTG/XP como "FIC" + tipo colados sem espaço.
    Sem isso, o has_token/has_any_phrase não reconhece o tipo do fundo, porque não há
    borda de palavra entre FIC e o restante (ex.: 'BTG Apolo II FICFIDC RL' não batia
    com FIDC nem com FIC isoladamente)."""
    return re.sub(r"\bFIC(FIDC|FIRF|FIA|FIM)\b", r"FIC \1", text)


def has_token(text: str, *terms: str) -> bool:
    """Procura termos como tokens, evitando COE dentro de PARTICIPACOES, FIA em CONFIANCA etc."""
    txt = norm(text)
    for term in terms:
        pattern = r"(?<![A-Z0-9])" + re.escape(norm(term)).replace(r"\ ", r"\s+") + r"(?![A-Z0-9])"
        if re.search(pattern, txt):
            return True
    return False


def has_any_phrase(text: str, terms: list[str] | tuple[str, ...]) -> bool:
    return any(has_token(text, term) for term in terms)


# Personalizações operacionais de liquidez. Chaves podem ser CNPJ (somente dígitos)
# ou trechos inequívocos do nome normalizado. 999 representa sem liquidez operacional />720 dias.
LIQUIDITY_OVERRIDES: dict[str, float] = {
    "CETIPADO": 999.0,
}


def liquidity_override_for_row(row: pd.Series) -> float:
    cnpj = only_digits_str(row.get("cnpj", ""))
    text = norm(" ".join([str(row.get("asset_id", "")), str(row.get("asset_nome", "")), str(row.get("manual_fundo", "")), str(row.get("manual_classe", ""))]))
    for key, days in LIQUIDITY_OVERRIDES.items():
        key_digits = only_digits_str(key)
        if key_digits and len(key_digits) >= 8 and cnpj == key_digits:
            return float(days)
        if not key_digits and has_token(text, key):
            return float(days)
    return np.nan


def format_brl_label(v) -> str:
    """Versão segura para labels markdown do Streamlit, evitando que o $ vire formatação."""
    return format_brl(v).replace("$", "\\$")


def format_brl(v) -> str:
    try:
        return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return "R$ 0,00"


def parse_brl_input(value, default: float = 0.0) -> float:
    """Converte campos digitados no padrão brasileiro, ex.: R$ 10.000,00."""
    try:
        s = str(value or "").replace("R$", "").replace(" ", "").strip()
        if not s:
            return default
        s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return default


def metric_card(label: str, value: str) -> None:
    html = (
        '<div class="mw-card">'
        f'<div class="mw-card-label">{escape(str(label))}</div>'
        f'<div class="mw-card-value">{escape(str(value))}</div>'
        '</div>'
    )
    st.markdown(html, unsafe_allow_html=True)


def page_intro(eyebrow: str, title: str, text: str, pills: list[str] | None = None) -> None:
    pill_html = "".join(f'<span class="mw-pill">{escape(str(p))}</span>' for p in (pills or []))
    st.markdown(
        '<div class="mw-page-intro">'
        f'<div class="mw-eyebrow">{escape(eyebrow)}</div>'
        f'<div class="mw-page-title">{escape(title)}</div>'
        f'<div class="mw-page-text">{escape(text)}</div>'
        + (f'<div style="margin-top:.65rem">{pill_html}</div>' if pill_html else '')
        + '</div>',
        unsafe_allow_html=True,
    )


def section_title(title: str, subtitle: str = "") -> None:
    st.markdown(f'<div class="mw-section-title">{escape(title)}</div>', unsafe_allow_html=True)
    if subtitle:
        st.markdown(f'<div class="mw-muted">{escape(subtitle)}</div>', unsafe_allow_html=True)


def ensure_saldo_operacional(df: pd.DataFrame) -> pd.DataFrame:
    """Garante identificação de saldo real mesmo quando o cache veio de versão antiga."""
    out = df.copy()
    if "saldo_operacional" not in out.columns:
        out["saldo_operacional"] = False
    base = out["saldo_operacional"].fillna(False).astype(bool)
    idx = out.index
    corretora = out.get("corretora", pd.Series([""] * len(out), index=idx)).astype(str).str.upper().str.strip()
    asset_id = out.get("asset_id", pd.Series([""] * len(out), index=idx)).astype(str).str.upper().str.strip()
    asset_nome = out.get("asset_nome", pd.Series([""] * len(out), index=idx)).astype(str).str.upper().str.strip()
    asset_tipo = out.get("asset_tipo", pd.Series([""] * len(out), index=idx)).astype(str).str.upper().str.strip()
    mercado = out.get("mercado", pd.Series([""] * len(out), index=idx)).astype(str).str.upper().str.strip()
    texto = asset_id + " " + asset_nome + " " + asset_tipo + " " + mercado
    xp_saldo = corretora.eq("XP") & asset_tipo.eq("FINANCEIRO") & asset_id.str.contains("SALDO", na=False)
    btg_saldo = corretora.eq("BTG") & (mercado.eq("CONTA CORRENTE") | asset_nome.eq("CONTA CORRENTE") | asset_id.eq("CONTA CORRENTE"))
    cs_saldo = corretora.eq("CS") & texto.str.contains("CASH|MONEY MARKET|SWEEP|BANK DEPOSIT", regex=True, na=False)
    out["saldo_operacional"] = (base | xp_saldo | btg_saldo | cs_saldo).fillna(False).astype(bool)
    return out


def friendly_class_name(x: str) -> str:
    mapa = {
        "RF Brasil": "Renda fixa no Brasil",
        "RV Brasil": "Renda variável no Brasil",
        "Internacional": "Investimentos internacionais",
        "Caixa": "Caixa",
        "Fora da Estratégia": "Fora da estratégia",
        "Alternativos": "Alternativos",
    }
    return mapa.get(str(x), str(x))


def friendly_strategy_name(x: str) -> str:
    mapa = {
        "Pós - Imediato": "Pós-fixado — liquidez imediata",
        "Pós - 1 a 30 dias": "Pós-fixado — resgate entre 1 e 30 dias",
        "Pós - 31 a 180 dias": "Pós-fixado — resgate entre 31 e 180 dias",
        "Pós - 181 a 360 dias": "Pós-fixado — resgate entre 181 e 360 dias",
        "Pós - 361+ dias": "Pós-fixado — resgate acima de 361 dias",
        "FiInfra Pós": "FI-Infra pós-fixado",
        "FiInfra Inflação": "FI-Infra indexado à inflação",
        "FiInfra e Cetipados": "Fundos de infraestrutura e crédito incentivado",
        "Pré - Bancário": "Prefixado bancário",
        "Pré - Tesouro": "Tesouro prefixado",
        "Inflação - Bancário": "Inflação bancário",
        "Inflação - Tesouro": "Tesouro indexado à inflação",
        "Crédito Privado": "Crédito privado",
        "Ações": "Ações brasileiras",
        "FIIs": "Fundos Imobiliários / FIAGROs",
        "Bitcoin": "Bitcoin",
        "Ouro": "Ouro",
        "Renda Fixa Internacional": "Renda fixa internacional",
        "Renda Variável Internacional": "Renda variável internacional",
        "Caixa Internacional": "Caixa internacional",
        "Saldo em Conta": "Saldo em conta",
        "Proventos a Receber": "Proventos a receber",
        "Direitos de Subscrição": "Direitos de subscrição",
        "Fundos de Investimento / Sem Liquidez Mapeada": "Fundos de investimento sem liquidez mapeada",
        "Previdência": "Previdência",
        "COE / Estruturados": "COE e estruturados",
        "Outros / Não Classificado": "Outros não classificados",
    }
    return mapa.get(str(x), str(x))


def money_color_styler(df: pd.DataFrame, money_cols=None, pct_cols=None, qty_cols=None, diff_cols=None):
    """Styler leve para tabelas pequenas, com diferença positiva em verde e negativa em vermelho."""
    money_cols = [c for c in (money_cols or []) if c in df.columns]
    pct_cols = [c for c in (pct_cols or []) if c in df.columns]
    qty_cols = [c for c in (qty_cols or []) if c in df.columns]
    diff_cols = [c for c in (diff_cols or []) if c in df.columns]

    fmt = {}
    for c in money_cols:
        fmt[c] = format_brl
    for c in pct_cols:
        fmt[c] = fmt_pct
    for c in qty_cols:
        fmt[c] = lambda x: "" if pd.isna(x) else f"{float(x):,.0f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def color_diff(v):
        try:
            val = float(v)
        except Exception:
            return ""
        if val > 0:
            return "color: #7bd88f; font-weight: 800;"
        if val < 0:
            return "color: #ff6b6b; font-weight: 800;"
        return "color: rgba(250,250,250,.72);"

    styler = df.style.format(fmt)
    if diff_cols:
        styler = styler.map(color_diff, subset=diff_cols)
    return styler


def macro_hierarchy_table(sub_df: pd.DataFrame, pl: float) -> pd.DataFrame:
    """Monta uma visão macro no estilo Power BI: classe principal + aberturas relevantes.

    A ideia é mostrar a carteira em uma leitura executiva, sem ficar presa só em
    Renda Fixa / Renda Variável / Internacional, mas também sem detalhar ativo a ativo.
    """
    if sub_df.empty:
        return pd.DataFrame(columns=[
            "Estratégia", "Quanto tem", "Quanto tem %", "Deveria ter (%)",
            "Deveria ter R$", "Ajuste necessário", "Diferença"
        ])

    base = sub_df.copy()
    for col in ["Valor Atual", "Valor Ideal", "Peso Atual", "Peso Ideal", "Diferença"]:
        if col in base.columns:
            base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0.0)

    def group_row(label: str, buckets: list[str] | None = None, classes: list[str] | None = None, is_header: bool = False):
        if classes is not None:
            part = base[base["Classe"].isin(classes)].copy()
        else:
            part = base[base["Subbucket"].isin(buckets or [])].copy()
        atual = float(part["Valor Atual"].sum()) if not part.empty else 0.0
        ideal = float(part["Valor Ideal"].sum()) if not part.empty else 0.0
        diff = ideal - atual
        return {
            "Estratégia": label,
            "Quanto tem": atual,
            "Quanto tem %": atual / pl if pl > 0 else 0.0,
            "Deveria ter (%)": ideal / pl if pl > 0 else 0.0,
            "Deveria ter R$": ideal,
            "Ajuste necessário": diff,
            "Diferença": (atual - ideal) / pl if pl > 0 else 0.0,
            "_header": is_header,
        }

    groups = [
        ("RENDA FIXA NO BRASIL", None, ["RF Brasil"], True),
        ("  Pós-fixado", ["Pós - Imediato", "Pós - 1 a 30 dias", "Pós - 31 a 180 dias", "Pós - 181 a 360 dias", "Pós - 361+ dias"], None, False),
        ("  Prefixado", ["Pré - Bancário", "Pré - Tesouro"], None, False),
        ("  Inflação", ["Inflação - Bancário", "Inflação - Tesouro"], None, False),
        ("  FI-Infra pós-fixado", ["FiInfra Pós"], None, False),
        ("  FI-Infra indexado à inflação", ["FiInfra Inflação"], None, False),
        ("  Crédito privado", ["Crédito Privado"], None, False),
        ("RENDA VARIÁVEL NO BRASIL", None, ["RV Brasil"], True),
        ("  Ações brasileiras", ["Ações"], None, False),
        ("  Fundos Imobiliários / FIAGROs", ["FIIs"], None, False),
        ("INVESTIMENTOS INTERNACIONAIS", None, ["Internacional"], True),
        ("  Renda fixa internacional", ["Renda Fixa Internacional"], None, False),
        ("  Renda variável internacional", ["Renda Variável Internacional"], None, False),
        ("ALTERNATIVOS", None, ["Alternativos"], True),
        ("  Bitcoin", ["Bitcoin"], None, False),
        ("  Ouro", ["Ouro"], None, False),
        ("FORA DA ESTRATÉGIA / MONITORAMENTO", None, ["Fora da Estratégia"], True),
        ("  Fundos sem liquidez mapeada", ["Fundos de Investimento / Sem Liquidez Mapeada"], None, False),
        ("  Previdência", ["Previdência"], None, False),
        ("  COE e estruturados", ["COE / Estruturados"], None, False),
        ("  Outros não classificados", ["Outros / Não Classificado"], None, False),
    ]

    rows = []
    for label, buckets, classes, is_header in groups:
        row = group_row(label, buckets=buckets, classes=classes, is_header=is_header)
        # Mostra sempre os cabeçalhos principais quando houver qualquer valor atual/ideal.
        # Nas linhas-filhas, remove ruído de zeros.
        if is_header:
            if abs(row["Quanto tem"]) > 0.01 or abs(row["Deveria ter R$"]) > 0.01:
                rows.append(row)
        else:
            if abs(row["Quanto tem"]) > 0.01 or abs(row["Deveria ter R$"]) > 0.01:
                rows.append(row)

    total_atual = float(base["Valor Atual"].sum())
    total_ideal = float(base["Valor Ideal"].sum())
    rows.append({
        "Estratégia": "TOTAL",
        "Quanto tem": total_atual,
        "Quanto tem %": total_atual / pl if pl > 0 else 0.0,
        "Deveria ter (%)": total_ideal / pl if pl > 0 else 0.0,
        "Deveria ter R$": total_ideal,
        "Ajuste necessário": total_ideal - total_atual,
        "Diferença": (total_atual - total_ideal) / pl if pl > 0 else 0.0,
        "_header": True,
    })
    return pd.DataFrame(rows)


def macro_hierarchy_styler(df: pd.DataFrame):
    """Aplica visual leve na tabela macro hierárquica."""
    visible_cols = ["Estratégia", "Quanto tem", "Quanto tem %", "Deveria ter (%)", "Deveria ter R$", "Ajuste necessário", "Diferença"]
    view = df[visible_cols].copy()
    header_mask = df.get("_header", pd.Series([False] * len(df))).astype(bool).tolist()

    fmt = {
        "Quanto tem": format_brl,
        "Quanto tem %": fmt_pct,
        "Deveria ter (%)": fmt_pct,
        "Deveria ter R$": format_brl,
        "Ajuste necessário": format_brl,
        "Diferença": fmt_pct,
    }

    def color_adjust(v):
        try:
            val = float(v)
        except Exception:
            return ""
        if val > 0:
            return "color: #7bd88f; font-weight: 800;"
        if val < 0:
            return "color: #ff6b6b; font-weight: 800;"
        return "color: rgba(250,250,250,.72);"

    def row_style(row):
        is_header = header_mask[row.name] if row.name < len(header_mask) else False
        if is_header:
            return [
                "background-color: rgba(93, 115, 170, .34); font-weight: 900; border-top: 1px solid rgba(255,255,255,.22);"
                for _ in row
            ]
        return ["" for _ in row]

    styler = view.style.format(fmt).apply(row_style, axis=1)
    styler = styler.map(color_adjust, subset=["Ajuste necessário"])
    return styler


def bucket_from_liquidity_days(days) -> str:
    try:
        d = float(days)
    except Exception:
        return "Fundos de Investimento / Sem Liquidez Mapeada"
    if d <= 1:
        return "Pós - Imediato"
    if d <= 30:
        return "Pós - 1 a 30 dias"
    if d <= 180:
        return "Pós - 31 a 180 dias"
    if d <= 360:
        return "Pós - 181 a 360 dias"
    return "Pós - 361+ dias"


def fund_name_key(value: str) -> str:
    """Chave robusta para casar fundos entre posição e Manual de Alocação.

    Remove pedaços jurídicos/comerciais que mudam entre corretoras, mas preserva
    os termos realmente distintivos do fundo/gestora.
    """
    s = norm(value)
    replacements = {
        "CRÉDITO PRIVADO": "CREDITO PRIVADO",
        "AÇÕES": "ACOES",
        "PREVIDÊNCIA": "PREVIDENCIA",
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    stop = [
        "FUNDO", "FUNDOS", "INVESTIMENTO", "INVESTIMENTOS", "COTAS", "COTA",
        "FIC", "FIF", "FIRF", "FIM", "FIA", "FIDC", "FI", "RF", "CP", "RL", "LP",
        "DE", "EM", "DO", "DA", "DAS", "DOS", "ADVISORY", "CREDITO", "PRIVADO",
        "PREVIDENCIA", "PGBL", "VGBL", "HIGH"  # HIGH fica pouco distintivo sozinho
    ]
    tokens = [t for t in re.split(r"\s+", s.strip()) if t and t not in stop]
    return " ".join(tokens)


def fund_token_set(value: str) -> set[str]:
    key = fund_name_key(value)
    return {t for t in key.split() if len(t) >= 3}


def fund_match_score(position_name: str, manual_name: str) -> float:
    """Score simples, sem dependência externa, para casar nomes parecidos.
    Prioriza sobreposição de tokens relevantes e sequência parcial.
    """
    a = fund_name_key(position_name)
    b = fund_name_key(manual_name)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.95
    ta, tb = set(a.split()), set(b.split())
    if not ta or not tb:
        return 0.0
    inter = ta & tb
    # mínimo de 2 tokens ajuda Jive BossaNova, BNP Rubi etc. sem casar coisas genéricas demais.
    token_score = len(inter) / max(1, min(len(ta), len(tb)))
    seq_score = 0.0
    try:
        from difflib import SequenceMatcher
        seq_score = SequenceMatcher(None, a, b).ratio()
    except Exception:
        seq_score = 0.0
    return max(token_score, seq_score * 0.82)


@st.cache_data(show_spinner=False)
def load_manual_fundos_cached(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> pd.DataFrame:
    """Lê o Manual de Alocação antigo, quando estiver disponível."""
    path = Path(path_str)
    if not path.exists():
        return pd.DataFrame(columns=["Fundo", "Classificação", "Liquidez (D+)", "Previdência", "CNPJ", "manual_key", "cnpj_norm"])
    try:
        raw = pd.read_excel(path, sheet_name="Gestoras e Fundos", header=None)
        header_row = None
        for i, row in raw.iterrows():
            vals = [str(x).strip() for x in row.tolist()]
            if "Gestora" in vals and "Fundo" in vals:
                header_row = i
                break
        if header_row is None:
            return pd.DataFrame(columns=["Fundo", "Classificação", "Liquidez (D+)", "Previdência", "CNPJ", "manual_key", "cnpj_norm"])
        df = pd.read_excel(path, sheet_name="Gestoras e Fundos", header=header_row)
        df.columns = [str(c).strip() for c in df.columns]
        keep = [c for c in ["Gestora", "Classificação", "Fundo", "Liquidez (D+)", "Perfil", "Condição", "Previdência", "Estretégia/Objetivo", "CNPJ"] if c in df.columns]
        df = df[keep].copy()
        df = df[df.get("Fundo", pd.Series(dtype=object)).notna()].copy()
        df["Fundo"] = df["Fundo"].astype(str).str.strip()
        if "CNPJ" not in df.columns:
            df["CNPJ"] = ""
        df["manual_key"] = df["Fundo"].apply(fund_name_key)
        df["cnpj_norm"] = df["CNPJ"].apply(only_digits_str)
        df["Liquidez (D+)"] = pd.to_numeric(df.get("Liquidez (D+)", np.nan), errors="coerce")
        df = df[df["manual_key"].str.len() > 2].drop_duplicates("manual_key")
        return df.reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=["Fundo", "Classificação", "Liquidez (D+)", "Previdência", "CNPJ", "manual_key", "cnpj_norm"])


def only_digits_str(value) -> str:
    raw = str(value or "").strip()
    # Excel frequentemente entrega CNPJ como float textual (ex.: 123...0).
    if re.fullmatch(r"\d+\.0", raw):
        raw = raw[:-2]
    digits = "".join(ch for ch in raw if ch.isdigit())
    return digits[:14] if len(digits) > 14 and raw.endswith(".0") else digits


def parse_days_from_text(value) -> float:
    """Extrai qualquer prazo de liquidez válido.

    A versão anterior aceitava apenas uma lista fechada de números. Assim, prazos
    reais como 44, 85, 119, 365 e 999 eram ignorados; quando a liquidação era D+1,
    o fundo acabava incorretamente no bucket de liquidez imediata.
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value) if float(value) >= 0 else np.nan
    txt = norm(value)
    if not txt or txt in ["NAN", "NONE", "NULL", "-"]:
        return np.nan
    m = re.search(r"D\s*\+\s*(\d+(?:[.,]\d+)?)", txt)
    if not m:
        # Aceita qualquer inteiro/decimal isolado, não apenas uma lista pré-definida.
        m = re.search(r"(?<![A-Z0-9])(\d+(?:[.,]\d+)?)(?![A-Z0-9])", txt)
    if not m:
        return np.nan
    try:
        return float(m.group(1).replace(",", "."))
    except Exception:
        return np.nan


def file_cache_signature(path: Path) -> tuple[int, int]:
    """Assinatura usada para invalidar cache quando uma planilha é substituída."""
    try:
        stt = path.stat()
        return int(stt.st_mtime_ns), int(stt.st_size)
    except Exception:
        return 0, 0


@st.cache_data(show_spinner=False)
def load_fundos_prev_cached(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> pd.DataFrame:
    """Lê a planilha Nome fundos e prev.xlsx, priorizando CNPJ, classificação e liquidez.

    Esta base é mais operacional para casar fundos da XP e BTG do que o manual antigo.
    """
    path = Path(path_str)
    cols = ["Fundo", "Classificação", "Liquidez (D+)", "Previdência", "CNPJ", "manual_key", "cnpj_norm", "Fonte", "Liquidez Operacional", "Classe Operacional", "Subbucket Operacional", "Rebalancear", "Status Mapeamento", "Observacao Operacional"]
    if not path.exists():
        return pd.DataFrame(columns=cols)
    rows = []
    try:
        xls = pd.ExcelFile(path)
        if "Fundos de Investimentos" in xls.sheet_names:
            df = pd.read_excel(path, sheet_name="Fundos de Investimentos")
            df.columns = [str(c).strip() for c in df.columns]
            df = canonicalize_master_columns(df, [
                "NOME_FUNDO", "CNPJ_FUNDO", "NOME_GESTORA", "CLASSIFICAÇÃO_CVM",
                "CLASSIFICAÇÃO_XP", "CAPTAÇÃO", "TIPO_INVESTIDOR",
                "MOVIMENTAÇÃO_MÍNIMA", "APLICAÇÃO_INICIAL_MÍNIMA",
                "COTIZAÇÃO_RESGATE", "PERÍODO_COTIZAÇÃO",
                "LIQUIDAÇÃO_RESGATE", "PERÍODO_LIQUIDAÇÃO",
                "CLASSE_OPERACIONAL", "SUBBUCKET_OPERACIONAL",
                "LIQUIDEZ_OPERACIONAL", "REBALANCEAR", "STATUS_MAPEAMENTO",
                "OBSERVACAO_OPERACIONAL",
            ])
            # Colunas opcionais permitem exceções no próprio arquivo mestre, sem editar código.
            # Exemplos: LIQUIDEZ_OPERACIONAL=999 e SUBBUCKET_OPERACIONAL="FiInfra Pós" ou "FiInfra Inflação".
            for _, r in df.iterrows():
                nome = str(r.get("NOME_FUNDO", "")).strip()
                if not nome or nome.lower() == "nan":
                    continue
                cot = parse_days_from_text(r.get("COTIZAÇÃO_RESGATE", np.nan))
                liq = parse_days_from_text(r.get("LIQUIDAÇÃO_RESGATE", np.nan))
                # Para alocação, o prazo de disponibilidade precisa considerar cotização + liquidação.
                prazo = np.nan
                if pd.notna(cot) or pd.notna(liq):
                    prazo = (0 if pd.isna(cot) else cot) + (0 if pd.isna(liq) else liq)
                rows.append({
                    "Fundo": nome,
                    "Classificação": str(r.get("CLASSIFICAÇÃO_XP", r.get("CLASSIFICAÇÃO_CVM", ""))).strip(),
                    "Classificação CVM": str(r.get("CLASSIFICAÇÃO_CVM", "")).strip(),
                    "Liquidez (D+)": prazo,
                    "Previdência": "Não",
                    "CNPJ": str(r.get("CNPJ_FUNDO", "")).strip(),
                    "manual_key": fund_name_key(nome),
                    "cnpj_norm": only_digits_str(r.get("CNPJ_FUNDO", "")),
                    "Fonte": "Cadastro Mestre - Fundos",
                    "Liquidez Operacional": parse_days_from_text(r.get("LIQUIDEZ_OPERACIONAL", np.nan)),
                    "Classe Operacional": str(r.get("CLASSE_OPERACIONAL", "")).strip(),
                    "Subbucket Operacional": str(r.get("SUBBUCKET_OPERACIONAL", "")).strip(),
                    "Rebalancear": str(r.get("REBALANCEAR", "")).strip(),
                    "Status Mapeamento": str(r.get("STATUS_MAPEAMENTO", "")).strip(),
                    "Observacao Operacional": str(r.get("OBSERVACAO_OPERACIONAL", "")).strip(),
                })
        if "Previdência" in xls.sheet_names:
            df = pd.read_excel(path, sheet_name="Previdência")
            df.columns = [str(c).strip() for c in df.columns]
            df = canonicalize_master_columns(df, [
                "Nome do Fundo Investido pelos planos", "CNPJ", "Status",
                "Liquidação ", "Liquidação", "Gestor Estratégico",
                "Classificação XP", "CLASSE_OPERACIONAL",
                "SUBBUCKET_OPERACIONAL", "LIQUIDEZ_OPERACIONAL",
                "REBALANCEAR", "STATUS_MAPEAMENTO",
                "OBSERVACAO_OPERACIONAL",
            ])
            for _, r in df.iterrows():
                nome = str(r.get("Nome do Fundo Investido pelos planos", "")).strip()
                if not nome or nome.lower() == "nan":
                    continue
                prazo = parse_days_from_text(r.get("Liquidação ", r.get("Liquidação", np.nan)))
                if pd.isna(prazo):
                    prazo = parse_days_from_text(r.get("Carência inicial para Port. Internas\n(dias corridos)", np.nan))
                rows.append({
                    "Fundo": nome,
                    "Classificação": str(r.get("Classificação XP", "")).strip(),
                    "Classificação CVM": "Previdência",
                    "Liquidez (D+)": prazo,
                    "Previdência": "Sim",
                    "CNPJ": str(r.get("CNPJ", "")).strip(),
                    "manual_key": fund_name_key(nome),
                    "cnpj_norm": only_digits_str(r.get("CNPJ", "")),
                    "Fonte": "Cadastro Mestre - Previdência",
                    "Liquidez Operacional": parse_days_from_text(r.get("LIQUIDEZ_OPERACIONAL", np.nan)),
                    "Classe Operacional": str(r.get("CLASSE_OPERACIONAL", "")).strip(),
                    "Subbucket Operacional": str(r.get("SUBBUCKET_OPERACIONAL", "")).strip(),
                    "Rebalancear": str(r.get("REBALANCEAR", "")).strip(),
                    "Status Mapeamento": str(r.get("STATUS_MAPEAMENTO", "")).strip(),
                    "Observacao Operacional": str(r.get("OBSERVACAO_OPERACIONAL", "")).strip(),
                })
    except Exception:
        pass
    out = pd.DataFrame(rows, columns=cols + ["Classificação CVM"])
    if not out.empty and "Liquidez Operacional" in out.columns:
        op = pd.to_numeric(out["Liquidez Operacional"], errors="coerce")
        out.loc[op.notna(), "Liquidez (D+)"] = op[op.notna()]
    if out.empty:
        return pd.DataFrame(columns=cols)
    out = out[out["manual_key"].astype(str).str.len() > 2].copy()
    # Prioriza CNPJ; mantém duplicatas de nome quando CNPJ diferente.
    out = out.drop_duplicates(subset=["cnpj_norm", "manual_key"], keep="first")
    return out.reset_index(drop=True)

def apply_manual_fund_mapping(df: pd.DataFrame) -> pd.DataFrame:
    """Casa fundos por hierarquia: CNPJ > nome exato > candidatos por tokens > fuzzy restrito.

    Evita o custo quadrático de comparar cada posição contra toda a base e mantém
    a base operacional Nome fundos e prev como fonte prioritária sobre o manual antigo.
    """
    out = df.copy()
    fundos_path = master_products_path()
    manual_path = find_file("Manual de Alocação.xlsx")
    fundos_sig = file_cache_signature(fundos_path)
    manual_sig = file_cache_signature(manual_path)
    base_nova = load_fundos_prev_cached(str(fundos_path), *fundos_sig)
    base_manual = load_manual_fundos_cached(str(manual_path), *manual_sig)
    bases = []
    if not base_nova.empty:
        bn = base_nova.copy(); bn["_priority"] = 0; bases.append(bn)
    if not base_manual.empty:
        bm = base_manual.copy(); bm["Fonte"] = "Manual de Alocação"; bm["_priority"] = 1
        if "Classificação CVM" not in bm.columns: bm["Classificação CVM"] = ""
        bases.append(bm)
    manual = pd.concat(bases, ignore_index=True, sort=False) if bases else pd.DataFrame()

    for col in ["manual_match", "manual_classe", "manual_liquidez", "manual_previdencia", "manual_fundo", "manual_cnpj", "manual_fonte", "manual_score", "manual_metodo", "manual_classe_operacional", "manual_subbucket_operacional", "manual_rebalancear", "manual_status_mapeamento", "manual_observacao_operacional"]:
        if col not in out.columns:
            out[col] = False if col == "manual_match" else (np.nan if col in ["manual_liquidez", "manual_score"] else "")
    if manual.empty:
        out["manual_match"] = False
        return out

    for c in ["asset_nome", "asset_id", "cnpj"]:
        if c not in out.columns: out[c] = ""
    out["_cnpj_norm"] = out["cnpj"].apply(only_digits_str)
    out["_fund_key_asset"] = out["asset_nome"].astype(str).apply(fund_name_key)
    empty_key = out["_fund_key_asset"].str.len().le(2)
    out.loc[empty_key, "_fund_key_asset"] = out.loc[empty_key, "asset_id"].astype(str).apply(fund_name_key)

    manual["cnpj_norm"] = manual.get("cnpj_norm", "").apply(only_digits_str)
    manual["manual_key"] = manual.get("manual_key", manual.get("Fundo", "")).astype(str)
    manual = manual.sort_values("_priority", kind="stable").drop_duplicates(["cnpj_norm", "manual_key"], keep="first")
    records = manual.to_dict("records")
    by_cnpj = {}
    by_key = {}
    token_index: dict[str, set[int]] = {}
    for i, rec in enumerate(records):
        cnpj = rec.get("cnpj_norm", "")
        key = str(rec.get("manual_key", ""))
        if len(cnpj) >= 8 and cnpj not in by_cnpj: by_cnpj[cnpj] = i
        if key and key not in by_key: by_key[key] = i
        for tok in set(key.split()):
            if len(tok) >= 4: token_index.setdefault(tok, set()).add(i)

    def find_match(cnpj: str, key: str, original_name: str = ""):
        cnpj = only_digits_str(cnpj)
        if cnpj and cnpj in by_cnpj: return records[by_cnpj[cnpj]], 1.0, "CNPJ"
        if key in by_key: return records[by_key[key]], 1.0, "Nome exato"
        tokens = {t for t in key.split() if len(t) >= 4}
        candidate_ids: set[int] = set()
        for tok in tokens: candidate_ids.update(token_index.get(tok, set()))
        if not candidate_ids: return None, 0.0, ""
        best_score, best_rec = 0.0, None
        source_name = original_name or key
        for i in candidate_ids:
            rec = records[i]
            score = fund_match_score(source_name, str(rec.get("Fundo", "")))
            if score > best_score: best_score, best_rec = score, rec
        return (best_rec, best_score, "Nome aproximado") if best_score >= 0.78 else (None, best_score, "")

    matches = [find_match(c, k, n) for c, k, n in zip(out["_cnpj_norm"], out["_fund_key_asset"], out["asset_nome"].astype(str))]
    recs = pd.Series([m[0] for m in matches], index=out.index)
    scores = pd.Series([m[1] for m in matches], index=out.index)
    methods = pd.Series([m[2] for m in matches], index=out.index)
    out["manual_match"] = recs.notna()
    mask = out["manual_match"]
    out.loc[mask, "manual_classe"] = recs[mask].apply(lambda r: str(r.get("Classificação", "")).strip())
    out.loc[mask, "manual_liquidez"] = recs[mask].apply(lambda r: pd.to_numeric(r.get("Liquidez (D+)", np.nan), errors="coerce"))
    out.loc[mask, "manual_previdencia"] = recs[mask].apply(lambda r: str(r.get("Previdência", "")).strip())
    out.loc[mask, "manual_fundo"] = recs[mask].apply(lambda r: str(r.get("Fundo", "")).strip())
    out.loc[mask, "manual_cnpj"] = recs[mask].apply(lambda r: only_digits_str(r.get("CNPJ", r.get("cnpj_norm", ""))))
    out.loc[mask, "manual_fonte"] = recs[mask].apply(lambda r: str(r.get("Fonte", "")).strip())
    out.loc[mask, "manual_classe_operacional"] = recs[mask].apply(lambda r: str(r.get("Classe Operacional", "")).strip())
    out.loc[mask, "manual_subbucket_operacional"] = recs[mask].apply(lambda r: str(r.get("Subbucket Operacional", "")).strip())
    out.loc[mask, "manual_rebalancear"] = recs[mask].apply(lambda r: str(r.get("Rebalancear", "")).strip())
    out.loc[mask, "manual_status_mapeamento"] = recs[mask].apply(lambda r: str(r.get("Status Mapeamento", "")).strip())
    out.loc[mask, "manual_observacao_operacional"] = recs[mask].apply(lambda r: str(r.get("Observacao Operacional", "")).strip())
    out.loc[mask, "manual_score"] = scores[mask]
    out.loc[mask, "manual_metodo"] = methods[mask]
    return out.drop(columns=["_fund_key_asset", "_cnpj_norm"], errors="ignore")


@st.cache_data(show_spinner=False)
def load_b3_master_cached(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> pd.DataFrame:
    """Lê o cadastro explícito por ticker da aba Ativos B3."""
    path = Path(path_str)
    cols = [
        "ticker_norm", "b3_nome", "b3_tipo_produto", "b3_classe_operacional",
        "b3_subbucket_operacional", "b3_estrategia", "b3_liquidez_operacional",
        "b3_fonte_preco", "b3_rebalancear", "b3_status_mapeamento",
        "b3_observacao_operacional", "b3_setor",
    ]
    if not path.exists():
        return pd.DataFrame(columns=cols)
    try:
        xls = pd.ExcelFile(path)
        if "Ativos B3" not in xls.sheet_names:
            return pd.DataFrame(columns=cols)
        raw = pd.read_excel(path, sheet_name="Ativos B3")
        raw.columns = [str(c).strip() for c in raw.columns]
        raw = canonicalize_master_columns(raw, [
            "TICKER", "NOME_ATIVO", "TIPO_PRODUTO", "CLASSE_OPERACIONAL",
            "SUBBUCKET_OPERACIONAL", "ESTRATEGIA", "REBALANCEAR",
            "STATUS_MAPEAMENTO", "OBSERVACAO_OPERACIONAL",
            "LIQUIDEZ_OPERACIONAL", "FONTE_PRECO", "SETOR",
        ])
        out = pd.DataFrame({
            "ticker_norm": raw.get("TICKER", "").astype(str).apply(ticker_clean),
            "b3_nome": raw.get("NOME_ATIVO", "").astype(str).str.strip(),
            "b3_tipo_produto": raw.get("TIPO_PRODUTO", "").astype(str).str.strip(),
            "b3_classe_operacional": raw.get("CLASSE_OPERACIONAL", "").astype(str).str.strip(),
            "b3_subbucket_operacional": raw.get("SUBBUCKET_OPERACIONAL", "").astype(str).str.strip(),
            "b3_estrategia": raw.get("ESTRATEGIA", "").astype(str).str.strip(),
            "b3_liquidez_operacional": pd.to_numeric(raw.get("LIQUIDEZ_OPERACIONAL", np.nan), errors="coerce"),
            "b3_fonte_preco": raw.get("FONTE_PRECO", "").astype(str).str.strip(),
            "b3_rebalancear": raw.get("REBALANCEAR", "").astype(str).str.strip(),
            "b3_status_mapeamento": raw.get("STATUS_MAPEAMENTO", "").astype(str).str.strip(),
            "b3_observacao_operacional": raw.get("OBSERVACAO_OPERACIONAL", "").astype(str).str.strip(),
            "b3_setor": raw.get("SETOR", "").astype(str).str.strip(),
        })
        out = out[out["ticker_norm"].apply(is_valid_b3_ticker)].copy()
        return out.drop_duplicates("ticker_norm", keep="first").reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=cols)


def apply_b3_master_mapping(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "ticker_norm" not in out.columns:
        out["ticker_norm"] = out.get("asset_id", "").astype(str).apply(ticker_clean)
    path = master_products_path()
    master = load_b3_master_cached(str(path), *file_cache_signature(path))
    if master.empty:
        out["b3_match"] = False
        return out
    out = out.merge(master, how="left", on="ticker_norm")
    out["b3_match"] = out["b3_classe_operacional"].fillna("").astype(str).str.len().gt(0) & out["b3_subbucket_operacional"].fillna("").astype(str).str.len().gt(0)
    return out



def price_source_for_row(row: pd.Series) -> str:
    """Resolve a fonte de preço operacional do cadastro mestre.

    Valores aceitos:
    - Yahoo Finance: ativo listado, remarcado por quantidade × cotação;
    - Valor da posição: mantém o valor informado pela corretora;
    - Não precificar: não tenta buscar cotação e não gera ordem por quantidade.

    Quando o campo estiver vazio, preserva o comportamento legado: fundos
    cetipados usam o valor da posição e os demais ativos elegíveis usam Yahoo.
    """
    fonte = norm(row.get("b3_fonte_preco", ""))
    tipo = norm(row.get("b3_tipo_produto", ""))
    if fonte in {"VALOR DA POSICAO", "VALOR POSICAO", "POSICAO", "CORRETORA"}:
        return "Valor da posição"
    if fonte in {"NAO PRECIFICAR", "SEM PRECIFICACAO", "NAO PRECIFICA"}:
        return "Não precificar"
    if fonte in {"YAHOO", "YAHOO FINANCE", "MERCADO"}:
        return "Yahoo Finance"
    if "CETIPADO" in tipo:
        return "Valor da posição"
    return "Yahoo Finance"


def exchange_position_mask(df: pd.DataFrame) -> pd.Series:
    """Posições listadas elegíveis para marcação a mercado pelo Yahoo.

    A posição continua sendo contabilizada mesmo quando REBALANCEAR=Não. Esse
    campo controla somente a geração da ordem. FONTE_PRECO controla a marcação.
    Custódia remunerada só entra na quantidade quando não existir a mesma ação
    na posição principal da mesma conta, evitando dupla contagem.
    """
    if df.empty:
        return pd.Series(dtype=bool, index=df.index)
    idx = df.index
    asset_tipo = df.get("asset_tipo", pd.Series("", index=idx)).astype(str).map(norm)
    mercado = df.get("mercado", pd.Series("", index=idx)).astype(str).map(norm)
    ticker = df.get("ticker_norm", pd.Series("", index=idx)).astype(str)
    classe = df.get("subbucket", pd.Series("", index=idx)).astype(str)
    fonte = df.get("fonte_preco", pd.Series("Yahoo Finance", index=idx)).astype(str).map(norm)
    tipo_cadastro = df.get("b3_tipo_produto", pd.Series("", index=idx)).astype(str).map(norm)

    real = (
        asset_tipo.isin(["ACOES", "FUNDOS IMOBILIARIOS", "CUSTODIA REMUNERADA"])
        | mercado.isin(["RENDA VARIAVEL", "RENDA VARIÁVEL", "CUSTODIA REMUNERADA"])
        | classe.isin(["Ações", "FIIs", "FiInfra Pós", "FiInfra Inflação", "FiInfra e Cetipados"])
    )
    excluded = asset_tipo.str.contains(
        "PROVENT|PROVISAO|EVENTO|OPCOES|OPÇÕES|OPCAO|OPÇÃO", regex=True, na=False
    )
    subscription_right = ticker.str.match(r"^[A-Z0-9]+12$", na=False)
    valid_ticker = ticker.apply(is_valid_b3_ticker)
    non_listed = tipo_cadastro.str.contains("CETIPADO|NAO LISTADO|NÃO LISTADO", regex=True, na=False)
    yahoo = fonte.eq("YAHOO FINANCE")

    # Evita somar a aba de aluguel quando a mesma conta/ticker já aparece na
    # posição principal. Quando só existe na custódia remunerada, ela é usada.
    custody = asset_tipo.eq("CUSTODIA REMUNERADA") | mercado.eq("CUSTODIA REMUNERADA")
    if custody.any() and "conta" in df.columns:
        keys = pd.DataFrame({
            "conta": df.get("conta", pd.Series("", index=idx)).astype(str),
            "ticker": ticker,
            "custody": custody,
        }, index=idx)
        regular_keys = set(map(tuple, keys.loc[~keys["custody"], ["conta", "ticker"]].values.tolist()))
        duplicated_custody = pd.Series(
            [(c, t) in regular_keys for c, t in keys[["conta", "ticker"]].itertuples(index=False, name=None)],
            index=idx,
        ) & custody
    else:
        duplicated_custody = pd.Series(False, index=idx)

    return (real & yahoo & ~non_listed & ~excluded & ~subscription_right & valid_ticker & ~duplicated_custody).fillna(False)


def ticker_rows(df: pd.DataFrame, ticker: str) -> pd.DataFrame:
    """Retorna linhas econômicas do ticker, incluindo Valor da posição."""
    tk = ticker_clean(ticker)
    if df.empty or not tk:
        return df.iloc[0:0].copy()
    rows = df[df.get("ticker_norm", pd.Series("", index=df.index)).astype(str).eq(tk)].copy()
    if rows.empty:
        return rows
    asset_tipo = rows.get("asset_tipo", pd.Series("", index=rows.index)).astype(str).map(norm)
    rows = rows[~asset_tipo.str.contains("PROVENT|PROVISAO|EVENTO|OPCOES|OPÇÕES|OPCAO|OPÇÃO", regex=True, na=False)]
    return rows


def yahoo_symbol(ticker: str) -> str:
    """Converte ticker B3 para o padrão do Yahoo Finance."""
    tk = ticker_clean(ticker)
    return f"{tk}.SA" if tk else ""


def _last_close_from_download(data: pd.DataFrame, yahoo_ticker: str) -> float:
    """Extrai o último Close válido de respostas simples ou MultiIndex."""
    if data is None or data.empty:
        return np.nan
    try:
        if isinstance(data.columns, pd.MultiIndex):
            # yfinance pode retornar (campo, ticker) ou (ticker, campo).
            if ("Close", yahoo_ticker) in data.columns:
                series = data[("Close", yahoo_ticker)]
            elif (yahoo_ticker, "Close") in data.columns:
                series = data[(yahoo_ticker, "Close")]
            else:
                candidates = [c for c in data.columns if "Close" in tuple(map(str, c)) and yahoo_ticker in tuple(map(str, c))]
                if not candidates:
                    return np.nan
                series = data[candidates[0]]
        else:
            if "Close" not in data.columns:
                return np.nan
            series = data["Close"]
        series = pd.to_numeric(series, errors="coerce").dropna()
        return float(series.iloc[-1]) if not series.empty else np.nan
    except Exception:
        return np.nan


@st.cache_data(ttl=300, show_spinner="Atualizando cotações de mercado...")
def load_yfinance_prices(tickers: tuple[str, ...]) -> dict[str, float]:
    """Busca preços recentes da B3 no Yahoo Finance.

    Primeiro tenta o último negócio intradiário de 1 minuto. Quando o Yahoo não
    disponibiliza intraday para algum ativo, utiliza o último fechamento diário.
    O cache de cinco minutos evita chamadas repetidas em cada interação do app.
    """
    clean = sorted({ticker_clean(t) for t in tickers if ticker_clean(t)})
    if not clean or not HAS_YFINANCE:
        return {}

    symbols = {tk: yahoo_symbol(tk) for tk in clean}
    yahoo_list = list(symbols.values())
    prices: dict[str, float] = {}

    try:
        intraday = yf.download(
            tickers=yahoo_list,
            period="1d",
            interval="1m",
            auto_adjust=False,
            progress=False,
            threads=True,
            group_by="column",
        )
        for tk, sym in symbols.items():
            px = _last_close_from_download(intraday, sym)
            if pd.notna(px) and px > 0:
                prices[tk] = float(px)
    except Exception:
        pass

    missing = [symbols[tk] for tk in clean if tk not in prices]
    if missing:
        try:
            daily = yf.download(
                tickers=missing,
                period="5d",
                interval="1d",
                auto_adjust=False,
                progress=False,
                threads=True,
                group_by="column",
            )
            reverse = {v: k for k, v in symbols.items()}
            for sym in missing:
                px = _last_close_from_download(daily, sym)
                if pd.notna(px) and px > 0:
                    prices[reverse[sym]] = float(px)
        except Exception:
            pass
    return prices


def mark_exchange_positions_to_market(df: pd.DataFrame, prices: dict[str, float]) -> pd.DataFrame:
    """Remarca posições reais de bolsa por quantidade × preço atual."""
    out = df.copy()
    mask = exchange_position_mask(out)
    if mask.empty or not mask.any():
        return out
    out.loc[mask, "quantidade"] = pd.to_numeric(out.loc[mask, "quantidade"], errors="coerce").fillna(0.0)
    live = out.loc[mask, "ticker_norm"].map(prices)
    valid = live.notna() & live.gt(0)
    idx_valid = live.index[valid]
    out.loc[idx_valid, "preco_mercado_atual"] = live.loc[idx_valid].astype(float)
    out.loc[idx_valid, "valor_mercado"] = (
        out.loc[idx_valid, "quantidade"].astype(float)
        * out.loc[idx_valid, "preco_mercado_atual"].astype(float)
    )
    return out


def current_exchange_position(pos_cliente: pd.DataFrame, ticker: str, price_ref: dict[str, float]) -> tuple[float, float, float]:
    """Preço, quantidade e valor atual respeitando a fonte de preço cadastrada."""
    tk = ticker_clean(ticker)
    rows = ticker_rows(pos_cliente, tk)
    if rows.empty:
        return float(price_ref.get(tk, np.nan)), 0.0, 0.0

    fonte = rows.get("fonte_preco", pd.Series("Yahoo Finance", index=rows.index)).astype(str).map(norm)
    yahoo_rows = rows[fonte.eq("YAHOO FINANCE")].copy()
    position_rows = rows[~fonte.eq("YAHOO FINANCE")].copy()
    preco = float(price_ref.get(tk, np.nan))
    qtd = float(pd.to_numeric(yahoo_rows.get("quantidade", 0.0), errors="coerce").fillna(0.0).sum()) if not yahoo_rows.empty else 0.0
    atual_yahoo = qtd * preco if pd.notna(preco) and preco > 0 else (float(pd.to_numeric(yahoo_rows.get("valor_mercado", 0.0), errors="coerce").fillna(0.0).sum()) if not yahoo_rows.empty else 0.0)
    atual_posicao = float(pd.to_numeric(position_rows.get("valor_mercado", 0.0), errors="coerce").fillna(0.0).sum()) if not position_rows.empty else 0.0
    return preco, qtd, atual_yahoo + atual_posicao


def ticker_can_rebalance(pos_cliente: pd.DataFrame, ticker: str) -> bool:
    rows = ticker_rows(pos_cliente, ticker)
    if rows.empty:
        return True
    return bool(rows.get("rebalancear", pd.Series(True, index=rows.index)).fillna(True).astype(bool).all())


def ticker_uses_quantity(pos_cliente: pd.DataFrame, ticker: str) -> bool:
    rows = ticker_rows(pos_cliente, ticker)
    if rows.empty:
        return True
    fonte = rows.get("fonte_preco", pd.Series("Yahoo Finance", index=rows.index)).astype(str).map(norm)
    return bool(fonte.eq("YAHOO FINANCE").all())


def operation_text(qtd_diff, diff_value) -> str:
    try:
        q = float(qtd_diff)
        v = float(diff_value)
    except Exception:
        return "Preço indisponível"
    if np.isnan(q):
        return "Preço indisponível"
    q_abs = abs(int(round(q)))
    if q_abs == 0:
        return "Manter"
    if v > 0:
        return f"Comprar {q_abs}"
    if v < 0:
        return f"Vender {q_abs}"
    return "Manter"



def basket_excel_bytes(df_orders: pd.DataFrame, cliente: str = "") -> BytesIO:
    """Gera um Excel simples de basket para acelerar a execução."""
    base = df_orders.copy()
    if base.empty or "Qtd a operar" not in base.columns:
        out = pd.DataFrame(columns=["C/V", "Ativo", "Quantidade", "Preço referência", "Valor estimado", "Cliente/Grupo"])
    else:
        base["Qtd a operar"] = pd.to_numeric(base["Qtd a operar"], errors="coerce").fillna(0)
        base = base[base["Qtd a operar"].round().ne(0)].copy()
        out = pd.DataFrame({
            "C/V": np.where(base["Qtd a operar"] > 0, "C", "V"),
            "Ativo": base["Ativo"].astype(str),
            "Quantidade": base["Qtd a operar"].abs().round().astype(int),
            "Preço referência": pd.to_numeric(base.get("Preço referência", np.nan), errors="coerce"),
            "Valor estimado": pd.to_numeric(base.get("Diferença", 0), errors="coerce").abs(),
            "Cliente/Grupo": cliente,
        })
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name="Basket")
    buf.seek(0)
    return buf


def dataframe_excel_bytes(df: pd.DataFrame, sheet_name: str = "Recomendações") -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    buf.seek(0)
    return buf


def format_usd(v) -> str:
    try:
        return f"US$ {float(v):,.2f}"
    except Exception:
        return "US$ 0.00"


def fmt_pct(x) -> str:
    try:
        return f"{100 * float(x):.2f}%".replace(".", ",")
    except Exception:
        return "0,00%"


def acao_por_diff(diff: float, tolerancia: float = 300.0) -> str:
    if diff > tolerancia:
        return "Comprar / Aportar"
    if diff < -tolerancia:
        return "Vender / Reduzir"
    return "Manter / OK"


def status_por_diff(diff: float, base: float, tolerancia_abs: float = 300.0, tolerancia_pct: float = 0.003) -> str:
    tol = max(tolerancia_abs, abs(base) * tolerancia_pct)
    if diff > tol:
        return "Abaixo do alvo"
    if diff < -tol:
        return "Acima do alvo"
    return "Dentro do alvo"


def prioridade_por_diff(diff: float, pl: float) -> str:
    if pl <= 0:
        return "Baixa"
    impacto = abs(diff) / pl
    if impacto >= 0.05:
        return "Alta"
    if impacto >= 0.015:
        return "Média"
    return "Baixa"


def prepare_display(
    df: pd.DataFrame,
    money_cols: list[str] | None = None,
    pct_cols: list[str] | None = None,
    qty_cols: list[str] | None = None,
    max_rows: int | None = None,
) -> pd.DataFrame:
    """Evita pandas Styler. Tabelas ficam mais leves e estáveis no Streamlit."""
    out = df.copy()
    if max_rows is not None:
        out = out.head(max_rows).copy()
    for c in money_cols or []:
        if c in out.columns:
            out[c] = out[c].apply(format_brl)
    for c in pct_cols or []:
        if c in out.columns:
            out[c] = out[c].apply(fmt_pct)
    for c in qty_cols or []:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).map(
                lambda x: f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
    return out


@st.cache_data(show_spinner=False)
def load_pesos_xlsx(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> dict[str, dict[str, float]]:
    path = Path(path_str)
    if not path.exists():
        return {}
    xls = pd.ExcelFile(path, engine="openpyxl")
    df = pd.read_excel(xls, sheet_name=xls.sheet_names[0], header=None).fillna("")
    pesos: dict[str, dict[str, float]] = {}
    carteira = None
    for _, row in df.iterrows():
        a = str(row.iloc[0]).strip()
        b = str(row.iloc[1]).strip() if len(row) > 1 else ""
        if not a and not b:
            continue
        if b.lower() == "neutro" and a:
            carteira = a
            pesos.setdefault(carteira, {})
            continue
        if carteira is None or not a:
            continue
        try:
            w = float(str(row.iloc[1]).replace(",", ".").strip())
        except Exception:
            w = 0.0
        # Mantém chaves iguais somadas, caso a planilha repita conceitos.
        pesos[carteira][a] = pesos[carteira].get(a, 0.0) + w
    return {k: v for k, v in pesos.items() if v}


@st.cache_data(show_spinner=False)
def load_contas_cached() -> pd.DataFrame:
    try:
        return posmod.load_control_accounts()
    except Exception:
        return pd.DataFrame()


@st.cache_data(show_spinner="Carregando base consolidada...")
def load_positions_cached(force_rebuild: bool = False) -> tuple[pd.DataFrame, dict, str]:
    """Abre imediatamente o último consolidado disponível.

    O rebuild completo ficou restrito ao botão de atualização ou à ausência do
    cache. Isso evita reler todos os arquivos XP/BTG/CS a cada inicialização —
    comportamento especialmente lento em deploys onde o mtime dos arquivos muda.
    """
    if force_rebuild:
        df = posmod.build_latest_from_repo()
        mode = "rebuild_manual"
    else:
        df = posmod.load_latest_positions()
        if df is None or df.empty:
            df = posmod.build_latest_from_repo()
            mode = "rebuild_empty"
        else:
            mode = "cache"
    meta = getattr(df, "attrs", {}).get("meta", {}) if hasattr(df, "attrs") else {}
    return df, meta, mode


# =============================================================================
# Classificação central dos ativos
# =============================================================================

def classify_fund_class(class_text: str, full_text: str, liquidez_days=np.nan) -> tuple[str, str]:
    """Converte a classificação de fundos para a estrutura operacional.

    Fundos de renda fixa, crédito e FIDC são distribuídos exclusivamente pelo
    prazo de liquidez. A natureza do fundo não deve contaminar o bucket de
    títulos de crédito privado, reservado a ativos diretos como debêntures,
    CRI, CRA e similares.
    """
    c = norm(class_text)
    t = norm(full_text)
    combined = expand_fund_compounds(f"{c} {t}")

    # Fundos tradicionais de infraestrutura (identificados por CNPJ) não são
    # confundidos com FI-Infra listados. Sem override operacional, seguem a
    # liquidez como qualquer outro fundo de renda fixa. Os FI-Infra listados são
    # classificados por ticker na função classify_position.

    # Internacional precisa ser identificado antes de palavras como AÇÕES/FIA.
    if any(x in combined for x in [
        "EXTERIOR", "INTERNACIONAL", "GLOBAL", "OFFSHORE", "DOLAR", "CAMBIAL", "USD"
    ]):
        if any(x in combined for x in ["BOND", "FIXED", "RENDA FIXA", "TREASURY", "CREDIT"]):
            return "Internacional", "Renda Fixa Internacional"
        return "Internacional", "Renda Variável Internacional"

    if has_any_phrase(combined, ["FII", "IMOBILIARIO", "REAL ESTATE"]):
        return "RV Brasil", "FIIs"
    if has_any_phrase(combined, ["ACOES", "RENDA VARIAVEL", "RV BRASIL", "SMALL CAPS", "IBOV", "LONG BIASED", "FIA"]):
        return "RV Brasil", "Ações"

    # FIDC, fundos de crédito, FIRF, DI, multimercado etc. entram pela liquidez.
    fund_rf_terms = [
        "FIDC", "DIREITOS CREDITORIOS", "CREDITO PRIVADO", "HIGH GRADE", "HIGH YIELD",
        "RENDA FIXA", "REFERENCIADO DI", "CDI", "FIRF", "MULTIMERCADO", "FIM",
        "MACRO", "RETORNO ABSOLUTO", "IPCA", "IMA-B", "PREFIXADO", "IRF-M"
    ]
    if has_any_phrase(combined, fund_rf_terms):
        return "RF Brasil", bucket_from_liquidity_days(liquidez_days)

    # Fundo sem classificação reconhecida: usa liquidez se disponível.
    if pd.notna(liquidez_days):
        return "RF Brasil", bucket_from_liquidity_days(liquidez_days)
    return "Fora da Estratégia", "Fundos de Investimento / Sem Liquidez Mapeada"


def infer_liquidity_days_from_row(row: pd.Series) -> float:
    """Hierarquia de liquidez: override > base manual > dados da corretora > nome.

    A base curada deve prevalecer sobre campos brutos da corretora, pois contempla
    exceções operacionais como fundos cetipados sem liquidez efetiva.
    """
    b3_operacional = pd.to_numeric(row.get("b3_liquidez_operacional", np.nan), errors="coerce")
    if pd.notna(b3_operacional): return float(b3_operacional)
    override = liquidity_override_for_row(row)
    if pd.notna(override): return float(override)
    manual = pd.to_numeric(row.get("manual_liquidez", np.nan), errors="coerce")
    if pd.notna(manual): return float(manual)
    cot = parse_days_from_text(row.get("cotizacao_resgate", ""))
    liq = parse_days_from_text(row.get("liquidacao_resgate", ""))
    if pd.notna(cot) or pd.notna(liq): return (0 if pd.isna(cot) else cot) + (0 if pd.isna(liq) else liq)
    direct = parse_days_from_text(row.get("liquidez", ""))
    if pd.notna(direct): return float(direct)
    for v in [row.get("asset_nome", ""), row.get("asset_id", "")]:
        d = parse_days_from_text(v)
        if pd.notna(d): return float(d)
    return np.nan

def classify_position(row: pd.Series) -> pd.Series:
    corretora = norm(row.get("corretora", ""))
    asset_id = ticker_clean(row.get("asset_id", ""))
    asset_tipo = norm(row.get("asset_tipo", ""))
    mercado = norm(row.get("mercado", ""))
    sub_mercado = norm(row.get("sub_mercado", ""))
    estrategia = norm(row.get("estrategia", ""))
    indexador = norm(row.get("indexador", ""))
    liquidez = norm(row.get("liquidez", ""))
    emissor = norm(row.get("emissor", ""))
    taxa = norm(row.get("taxa", ""))
    nome = norm(row.get("asset_nome", ""))
    cnpj = only_digits_str(row.get("cnpj", ""))
    text = expand_fund_compounds(" ".join([asset_id, nome, asset_tipo, mercado, sub_mercado, estrategia, indexador, liquidez, emissor, taxa]))
    # Flags robustas de indexador. Alguns relatórios trazem prefixado como
    # "CDB PRE DU", "CRA PRE DU" ou apenas "PRE" no nome, sem escrever
    # "prefixado". Usamos regex com borda de palavra para não confundir com
    # PREV/PREVIDÊNCIA.
    has_ipca = bool(re.search(r"\b(IPCA|IPC\s*-?\s*A|INFLACAO|INFLAÇÃO)\b", text))
    has_cdi = bool(re.search(r"\b(CDI|DI|POS\s*-?\s*FIXADO|PÓS\s*-?\s*FIXADO|POS\s+FIXADO|PÓS\s+FIXADO)\b", text))
    has_pre = bool(re.search(r"\b(PRE|PRÉ|PREFIXADO)\b|PRE\s*-\s*FIXADO|PRÉ\s*-\s*FIXADO|PRE\s+FIXADO|PRÉ\s+FIXADO", text))

    # Sinais vindos diretamente dos relatórios.
    # XP: no relatório de Renda Fixa, prefixados normalmente vêm com NomeIndexador vazio
    # e TaxaCompleta como '+13,10%' / '+14,50%', enquanto CDI/IPCA aparecem no NomeIndexador
    # ou na TaxaCompleta.
    # BTG: a coluna Estratégia costuma trazer Pré-fixado/Pós-fixado/Inflação; quando vem
    # em branco, a Taxa Compra/Taxa Emissão ajuda no fallback.
    indexador_blank = indexador in ["", "NAN", "NONE", "NULL", "-", "0"]
    estrategia_blank = estrategia in ["", "NAN", "NONE", "NULL", "-"]
    taxa_clean = taxa.replace("%", "").replace("+", "").replace(",", ".").strip()
    try:
        taxa_num = float(re.sub(r"[^0-9.\-]", "", taxa_clean)) if taxa_clean else np.nan
    except Exception:
        taxa_num = np.nan
    has_taxa_nominal = pd.notna(taxa_num) and abs(float(taxa_num)) > 0.0001 and not has_cdi and not has_ipca
    pre_by_report = (has_pre or (indexador_blank and has_taxa_nominal and ("RENDA FIXA" in mercado or "RENDA FIXA" in asset_tipo or asset_id.startswith(("CDB", "LCI", "LCA", "LCD", "LF", "DEB", "CRI", "CRA", "CDCA")))))
    liq_days = infer_liquidity_days_from_row(row)

    # Cadastro explícito por ticker tem prioridade sobre a base de fundos.
    # Isso evita conflito quando um código listado também aparece em cadastros
    # legados de fundos.
    if bool(row.get("b3_match", False)):
        b3_classe = optional_text(row.get("b3_classe_operacional", ""))
        b3_bucket = optional_text(row.get("b3_subbucket_operacional", ""))
        b3_tipo = optional_text(row.get("b3_tipo_produto", ""))
        if b3_bucket == "FiInfra e Cetipados" and asset_id in FI_INFRA_TICKERS:
            b3_bucket = "FiInfra Pós" if asset_id in FI_INFRA_POS_TICKERS else "FiInfra Inflação"
        return pd.Series([b3_classe, b3_bucket, b3_bucket, f"Cadastro Mestre B3{(' / ' + b3_tipo) if b3_tipo else ''}"])

    # Exceção operacional por CNPJ/nome na base de fundos.
    op_classe = optional_text(row.get("manual_classe_operacional", ""))
    op_bucket = optional_text(row.get("manual_subbucket_operacional", ""))
    if op_classe and op_bucket:
        return pd.Series([op_classe, op_bucket, op_bucket, "Override operacional / base de fundos"])

    # Eventos não são ativos fora da estratégia. Proventos ficam como caixa a
    # receber e códigos final 12 são direitos de subscrição monitorados em RV.
    if asset_tipo in ["PROVENTOS", "PROVENTOS FUNDO IMOB"] or "PROVENTO" in asset_tipo:
        return pd.Series(["Caixa", "Proventos a Receber", "Proventos a Receber", "Evento / não rebalancear"])
    # BTG: dividendos, JCP e contribuições de previdência aparecem na aba "Valor em
    # Trânsito" enquanto aguardam liquidação. Sem esta regra caíam em "Outros / Não
    # Classificado" e ficavam de fora do casamento de estratégia.
    if mercado == "VALOR EM TRANSITO":
        return pd.Series(["Caixa", "Proventos a Receber", "Proventos a Receber", "Evento / não rebalancear / BTG Valor em Trânsito"])
    if re.fullmatch(r"[A-Z]{4}12", asset_id):
        return pd.Series(["RV Brasil", "Direitos de Subscrição", "Direitos de Subscrição", "Direito / não rebalancear"])

    if bool(row.get("saldo_operacional", False)):
        if corretora == "CS":
            return pd.Series(["Internacional", "Caixa Internacional", "Caixa Internacional", "Operacional"])
        return pd.Series(["Caixa", "Saldo em Conta", "Saldo em Conta", "Operacional"])

    # A natureza econômica prevalece inclusive quando a corretora entrega o
    # ativo em uma aba incorreta, como debênture dentro de Renda Variável.
    early_is_fiinfra = asset_id in FI_INFRA_TICKERS
    if early_is_fiinfra:
        bucket_infra = "FiInfra Pós" if asset_id in FI_INFRA_POS_TICKERS else "FiInfra Inflação"
        return pd.Series(["RF Brasil", bucket_infra, bucket_infra, "Natureza econômica / FI-Infra listado"])

    # Linhas vindas da aba de Fundos Imobiliários da corretora são posições
    # listadas. Essa regra vem antes da heurística de fundos, pois o texto
    # "Fundos Imobiliários" contém a palavra FUNDOS e, sem esta proteção,
    # FIAGROs/FIPs como KNCA11, RZAG11 e VIGT11 caíam em fundo sem liquidez.
    if asset_tipo in ["FUNDOS IMOBILIARIOS", "FUNDOS IMOBILIÁRIOS"] or mercado in ["FUNDOS IMOBILIARIOS", "FUNDOS IMOBILIÁRIOS"]:
        return pd.Series(["RV Brasil", "FIIs", "FIIs", "Aba de Fundos Imobiliários / fallback"])

    # Ações emprestadas via "Custódia Remunerada" (XP) continuam sendo a mesma ação
    # do cliente, só que remunerada; a aba não é reconhecida por nenhuma outra regra.
    if asset_tipo == "CUSTODIA REMUNERADA" or mercado == "CUSTODIA REMUNERADA":
        if re.fullmatch(r"[A-Z0-9]{4}\d{1,2}", asset_id):
            return pd.Series(["RV Brasil", "Ações", "Ações", "Custódia remunerada / aluguel de ações"])

    early_looks_like_fund = has_any_phrase(text, ["FIC", "FIM", "FIRF", "FIA", "FIDC", "FUNDO", "FUNDOS", "FIF"])
    early_credit_title = bool(re.search(r"\b(DEB|DEBENTURE|DEBENTURES|CRI|CRA|CDCA)\b", text))
    if early_credit_title and not early_looks_like_fund:
        return pd.Series(["RF Brasil", "Crédito Privado", "Crédito Privado", "Natureza econômica / Crédito Privado"])

    # Fundo também prevalece sobre a aba de origem. Isso corrige fundos
    # internacionais e FIDC eventualmente entregues na aba de ações.
    if early_looks_like_fund:
        classe_base = row.get("manual_classe", "") if row.get("manual_match", False) else text
        classe, bucket = classify_fund_class(classe_base, text, liq_days)
        origem = "Manual/Fundos" if row.get("manual_match", False) else "Heurística de fundo"
        if asset_tipo in ["PREVIDENCIA", "PREVIDÊNCIA"] or any(x in text for x in ["PREVIDENCIA", "PGBL", "VGBL"]):
            origem = "Previdência"
        return pd.Series([classe, bucket, bucket, origem])

    # BTG: Mercado/Sub Mercado/Estratégia normalmente já dizem exatamente o direcionamento.
    if corretora == "BTG":
        if mercado == "CONTA CORRENTE" or sub_mercado == "CC":
            return pd.Series(["Caixa", "Saldo em Conta", "Saldo em Conta", "Operacional"])
        if mercado == "PREVIDENCIA" or sub_mercado == "PP":
            classe, bucket = classify_fund_class(row.get("manual_classe", "") or estrategia, text, liq_days)
            return pd.Series([classe, bucket, bucket, "Previdência"])
        if mercado == "RENDA FIXA":
            if sub_mercado in ["TESOURO DIRETO", "TITULO PUBLICO", "TÍTULO PÚBLICO"]:
                if any(x in estrategia + " " + text for x in ["INFLACAO", "INFLAÇÃO", "IPCA", "NTNB", "NTN-B"]):
                    return pd.Series(["RF Brasil", "Inflação - Tesouro", "Inflação - Tesouro", "BTG"])
                if any(x in estrategia + " " + text for x in ["PRE", "PRÉ", "PREFIXADO", "LTN", "NTN-F"]):
                    return pd.Series(["RF Brasil", "Pré - Tesouro", "Pré - Tesouro", "BTG"])
                return pd.Series(["RF Brasil", "Pós - Imediato", "Pós - Imediato", "BTG"])
            # Fora de títulos bancários e públicos, toda renda fixa direta é crédito privado.
            bank_title = bool(re.match(r"^(CDB|LCI|LCA|LCD|LF)\b", asset_id)) or any(
                x in sub_mercado for x in ["BANCARIO", "BANCÁRIO"]
            )
            if not bank_title:
                return pd.Series(["RF Brasil", "Crédito Privado", "Crédito Privado", "BTG / título de crédito privado"])
            if "INFL" in estrategia or has_ipca:
                return pd.Series(["RF Brasil", "Inflação - Bancário", "Inflação - Bancário", "BTG Estratégia/Indexador"])
            if "PRE" in estrategia or "PRÉ" in estrategia or pre_by_report or (estrategia_blank and has_taxa_nominal):
                return pd.Series(["RF Brasil", "Pré - Bancário", "Pré - Bancário", "BTG Estratégia/Taxa"])
            return pd.Series(["RF Brasil", "Pós - 361+ dias", "Pós - 361+ dias", "BTG Estratégia/Indexador"])
        if mercado == "FUNDOS":
            classe, bucket = classify_fund_class(row.get("manual_classe", "") or estrategia, text, liq_days)
            return pd.Series([classe, bucket, bucket, "Manual" if row.get("manual_match", False) else "BTG/Heurística"])
        if mercado == "RENDA VARIAVEL" or mercado == "RENDA VARIÁVEL":
            if asset_id in ATIVOS_ESTRATEGICOS_B3:
                info = ATIVOS_ESTRATEGICOS_B3[asset_id]
                return pd.Series([info["classe"], info["subbucket"], info["subbucket"], f"ETF B3: {info['estrategia']}"])
            if sub_mercado == "FII" or asset_id.endswith("11"):
                # FiInfra antes de FII comum.
                if asset_id in FI_INFRA_TICKERS:
                    bucket_infra = "FiInfra Pós" if asset_id in FI_INFRA_POS_TICKERS else "FiInfra Inflação"
                    return pd.Series(["RF Brasil", bucket_infra, bucket_infra, "FI-Infra listado"])
                return pd.Series(["RV Brasil", "FIIs", "FIIs", "Estratégia"])
            if sub_mercado == "ACAO" or len(asset_id) in [5, 6]:
                return pd.Series(["RV Brasil", "Ações", "Ações", "Estratégia"])
        if mercado == "COE":
            return pd.Series(["Fora da Estratégia", "COE / Estruturados", "COE / Estruturados", "Fora da Estratégia"])

    # Natureza econômica prevalece sobre a aba de origem do relatório.
    # Isso corrige ativos que chegam em abas inadequadas (ex.: debênture em Ações)
    # e fundos de infraestrutura cujo nome também contém FIA/FII/ações.
    is_fiinfra = asset_id in FI_INFRA_TICKERS
    if is_fiinfra:
        bucket_infra = "FiInfra Pós" if asset_id in FI_INFRA_POS_TICKERS else "FiInfra Inflação"
        return pd.Series(["RF Brasil", bucket_infra, bucket_infra, "Natureza econômica / FI-Infra listado"])

    is_credit_title = bool(re.search(r"\b(DEB|DEBENTURE|DEBENTURES|DEBÊNTURE|DEBÊNTURES|CRI|CRA|CDCA)\b", text))
    looks_like_fund = has_any_phrase(text, ["FIC", "FIM", "FIRF", "FIA", "FIDC", "FUNDO", "FUNDOS", "FIF"])
    if is_credit_title and not looks_like_fund:
        return pd.Series(["RF Brasil", "Crédito Privado", "Crédito Privado", "Natureza econômica / Crédito Privado"])

    # XP: fundos e previdência têm CNPJ/prazos na própria planilha e/ou na base Nome fundos e prev.
    if asset_tipo in ["FUNDOS", "PREVIDENCIA", "PREVIDÊNCIA"] or has_any_phrase(text, ["FIC", "FIM", "FIRF", "FIA", "FIDC", "FUNDO", "FUNDOS", "FIF"]):
        classe_base = row.get("manual_classe", "") if row.get("manual_match", False) else ""
        if not classe_base:
            classe_base = text
        classe, bucket = classify_fund_class(classe_base, text, liq_days)
        origem = "Manual/Fundos" if row.get("manual_match", False) else "XP/Heurística"
        if asset_tipo in ["PREVIDENCIA", "PREVIDÊNCIA"] or any(x in text for x in ["PREV", "PREVIDENCIA", "PREVIDÊNCIA", "PGBL", "VGBL"]):
            # Mantém o recurso dentro da classe econômica do fundo, mas marca a origem como previdência.
            return pd.Series([classe, bucket, bucket, "Previdência"])
        return pd.Series([classe, bucket, bucket, origem])

    # Internacional
    if corretora == "CS":
        if any(x in text for x in ["CASH", "MONEY MARKET", "SWEEP", "BANK DEPOSIT"]):
            return pd.Series(["Internacional", "Caixa Internacional", "Caixa Internacional", "Operacional"])
        if any(x in text for x in ["BOND", "FIXED", "TREASURY", "CD ", "CERTIFICATE", "NOTE", "CORPORATE"]):
            return pd.Series(["Internacional", "Renda Fixa Internacional", "Renda Fixa Internacional", "Estratégia"])
        return pd.Series(["Internacional", "Renda Variável Internacional", "Renda Variável Internacional", "Estratégia"])

    # Caixa / saldo fallback conservador
    if any(x in text for x in ["SALDO FINANCEIRO", "CONTA CORRENTE", "VALORDISPONIVEL"]):
        return pd.Series(["Caixa", "Saldo em Conta", "Saldo em Conta", "Operacional"])

    # COE / estruturados
    if has_any_phrase(text, ["COE", "ESTRUTURADO", "OPCOES FLEX", "OPCAO FLEX", "OPCOES", "OPÇÃO"]):
        return pd.Series(["Fora da Estratégia", "COE / Estruturados", "COE / Estruturados", "Fora da Estratégia"])

    # Ativos estratégicos negociados na B3 que não podem cair na regra genérica de FII.
    if asset_id in ATIVOS_ESTRATEGICOS_B3:
        info = ATIVOS_ESTRATEGICOS_B3[asset_id]
        return pd.Series([info["classe"], info["subbucket"], info["subbucket"], f"ETF B3: {info['estrategia']}"])

    # Fi-Infra: antes de FII, porque todos terminam em 11.
    if asset_id in FI_INFRA_TICKERS:
        bucket_infra = "FiInfra Pós" if asset_id in FI_INFRA_POS_TICKERS else "FiInfra Inflação"
        return pd.Series(["RF Brasil", bucket_infra, bucket_infra, "FI-Infra listado"])

    # Bolsa Brasil
    if asset_tipo in ["FUNDOS IMOBILIARIOS", "FUNDOS IMOBILIÁRIOS"] or has_any_phrase(text, ["FUNDO IMOB", "FII", "FUNDO IMOBILIARIO"]):
        return pd.Series(["RV Brasil", "FIIs", "FIIs", "Estratégia"])
    if (asset_id.endswith("11") and len(asset_id) >= 5 and asset_id[:4].isalpha()):
        return pd.Series(["RV Brasil", "FIIs", "FIIs", "Estratégia"])
    if asset_tipo in ["ACOES", "AÇÕES"] or has_any_phrase(text, ["ACAO", "AÇÃO", "BOVESPA", "RENDA VARIAVEL"]):
        return pd.Series(["RV Brasil", "Ações", "Ações", "Estratégia"])
    if len(asset_id) in [5, 6] and asset_id[:4].isalpha() and asset_id[-1].isdigit():
        return pd.Series(["RV Brasil", "Ações", "Ações", "Estratégia"])

    # Tesouro / títulos públicos são Tesouro na estratégia.
    if any(x in text for x in ["TESOURO SELIC", "LFT", "SELIC"]):
        return pd.Series(["RF Brasil", "Pós - Imediato", "Pós - Imediato", "Estratégia"])
    if any(x in text for x in ["TESOURO PRE", "NTN-F", "NTNF", "LTN"]):
        return pd.Series(["RF Brasil", "Pré - Tesouro", "Pré - Tesouro", "Estratégia"])
    if any(x in text for x in ["TESOURO IPCA", "NTN-B", "NTNB", "NTNB PRINC"]):
        return pd.Series(["RF Brasil", "Inflação - Tesouro", "Inflação - Tesouro", "Estratégia"])

    # Renda fixa local por indexador/taxa.
    is_rf_local = ("RENDA FIXA" in mercado or "RENDA FIXA" in asset_tipo or asset_id.startswith(("CDB", "LCI", "LCA", "LCD", "LF", "DEB", "CRI", "CRA", "CDCA")))
    if is_rf_local or has_any_phrase(text, ["CDB", "LCI", "LCA", "LCD", "LF", "CRI", "CRA", "DEB", "DEBENTURE", "CDCA"]):
        bank_title = asset_id.startswith(("CDB", "LCI", "LCA", "LCD", "LF"))
        if not bank_title:
            return pd.Series(["RF Brasil", "Crédito Privado", "Crédito Privado", "Título privado não bancário"])
        if has_ipca:
            return pd.Series(["RF Brasil", "Inflação - Bancário", "Inflação - Bancário", "Relatório: indexador/taxa"])
        if pre_by_report:
            return pd.Series(["RF Brasil", "Pré - Bancário", "Pré - Bancário", "Relatório: indexador/taxa"])
        return pd.Series(["RF Brasil", "Pós - 361+ dias", "Pós - 361+ dias", "Relatório: indexador/taxa"])

    return pd.Series(["Fora da Estratégia", "Outros / Não Classificado", "Outros / Não Classificado", "Revisar"])
@st.cache_data(show_spinner=False)
def enrich_positions_cached(df: pd.DataFrame, mapping_signature: tuple = ()) -> pd.DataFrame:
    df = df.copy()
    df = df.loc[:, ~df.columns.duplicated()].copy()
    df = df.drop(columns=[c for c in ["classe_macro", "subclasse", "subbucket", "tratamento"] if c in df.columns], errors="ignore")
    for c in ["valor_mercado", "quantidade", "valor_original"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["ticker_norm"] = df.get("asset_id", "").astype(str).apply(ticker_clean)
    df = apply_manual_fund_mapping(df)
    df = apply_b3_master_mapping(df)

    # Consolida os campos operacionais para auditoria e uso no motor.
    df["fonte_preco"] = df.apply(price_source_for_row, axis=1)
    df["liquidez_operacional_aplicada"] = df.apply(infer_liquidity_days_from_row, axis=1)

    # REBALANCEAR do cadastro mestre é aplicado sem apagar o comportamento padrão.
    manual_flag = df.get("manual_rebalancear", pd.Series("", index=df.index)).astype(str)
    b3_flag = df.get("b3_rebalancear", pd.Series("", index=df.index)).astype(str)
    df["rebalancear"] = True
    has_manual_flag = manual_flag.map(norm).isin(["SIM", "NAO", "NÃO", "S", "N", "TRUE", "FALSE", "1", "0"])
    has_b3_flag = b3_flag.map(norm).isin(["SIM", "NAO", "NÃO", "S", "N", "TRUE", "FALSE", "1", "0"])
    if has_manual_flag.any():
        df.loc[has_manual_flag, "rebalancear"] = manual_flag.loc[has_manual_flag].map(parse_yes_no).astype(bool).to_numpy()
    if has_b3_flag.any():
        df.loc[has_b3_flag, "rebalancear"] = b3_flag.loc[has_b3_flag].map(parse_yes_no).astype(bool).to_numpy()

    cols = df.apply(classify_position, axis=1)
    cols.columns = ["classe_macro", "subclasse", "subbucket", "tratamento"]
    out = pd.concat([df, cols], axis=1)
    out = out.loc[:, ~out.columns.duplicated()].copy()
    return out


# =============================================================================
# Modelo/targets
# =============================================================================
def peso_get(p: dict[str, float], key: str) -> float:
    key_n = norm(key)
    for k, v in p.items():
        if norm(k) == key_n:
            return float(v or 0)
    return 0.0


def model_for_profile(perfil: str, modelos: list[str]) -> str:
    pn = norm(perfil)
    candidates = []
    if "ARROJADO RENDA CONSTRUCAO" in pn: candidates.append("Arrojado Renda Construção")
    if "MODERADO RENDA CONSTRUCAO" in pn: candidates.append("Moderado Renda Construção")
    if "CONSERVADOR RENDA CONSTRUCAO" in pn: candidates.append("Conservador Renda Construção")
    if "ARROJADO RENDA USUFRUTO" in pn: candidates.append("Arrojado Renda Usufruto")
    if "MODERADO RENDA USUFRUTO" in pn: candidates.append("Moderado Renda Usufruto")
    if "CONSERVADOR RENDA USUFRUTO" in pn: candidates.append("Conservador Renda Usufruto")
    if "ULTRACONSERVADOR" in pn: candidates.append("Ultraconservador")
    if "CONSERVADOR" in pn: candidates.append("Conservador")
    if "MODERADO" in pn: candidates.append("Moderado")
    if "ARROJADO" in pn: candidates.append("Arrojado")
    for c in candidates:
        for m in modelos:
            if norm(m) == norm(c):
                return m
    return modelos[0] if modelos else ""


def subbucket_targets_from_model(p: dict[str, float], pl: float) -> pd.DataFrame:
    rows = [
        ("RF Brasil", "Pós - Imediato", peso_get(p, "Imediato")),
        ("RF Brasil", "Pós - 1 a 30 dias", peso_get(p, "1 a 30 dias")),
        ("RF Brasil", "Pós - 31 a 180 dias", peso_get(p, "31 a 180 dias")),
        ("RF Brasil", "Pós - 181 a 360 dias", peso_get(p, "181 a 360 dias")),
        ("RF Brasil", "Pós - 361+ dias", peso_get(p, "361+ dias")),
        ("RF Brasil", "FiInfra Pós", peso_get(p, "FiInfra e Cetipados")),
        ("RF Brasil", "FiInfra Inflação", peso_get(p, "FiInfra e Cetipado")),
        ("RF Brasil", "Pré - Bancário", peso_get(p, "Bancário Pré")),
        ("RF Brasil", "Pré - Tesouro", peso_get(p, "Tesouro Pré")),
        ("RF Brasil", "Inflação - Bancário", peso_get(p, "Bancário")),
        ("RF Brasil", "Inflação - Tesouro", peso_get(p, "Tesouro")),
        ("RF Brasil", "Crédito Privado", peso_get(p, "Crédito Privado")),
        ("RV Brasil", "Ações", peso_get(p, "Ações")),
        ("RV Brasil", "FIIs", peso_get(p, "FIIs")),
        ("Internacional", "Renda Fixa Internacional", peso_get(p, "Renda Fixa")),
        ("Internacional", "Renda Variável Internacional", peso_get(p, "Renda Variável")),
    ]
    df = pd.DataFrame(rows, columns=["Classe", "Subbucket", "Peso Ideal"])
    df = df[df["Peso Ideal"] > 0].copy()
    df["Valor Ideal"] = df["Peso Ideal"] * pl
    return df


def macro_targets_from_sub(subtarget: pd.DataFrame, pl: float, p: dict[str, float]) -> pd.DataFrame:
    if subtarget.empty:
        rv = peso_get(p, "RV Brasil")
        intl = peso_get(p, "Internacional")
        rf = max(0.0, 1.0 - rv - intl)
        return pd.DataFrame({"Classe": ["RF Brasil", "RV Brasil", "Internacional"], "Peso Ideal": [rf, rv, intl], "Valor Ideal": [rf*pl, rv*pl, intl*pl]})
    macro = subtarget.groupby("Classe", as_index=False).agg({"Peso Ideal": "sum", "Valor Ideal": "sum"})
    return macro


def rv_universe(modelo: str) -> dict[str, list[str]]:
    return {"Ações": ACOES_COM_RENDA if "RENDA" in norm(modelo) else ACOES_SEM_RENDA, "FIIs": FIIS_RECOMENDADOS}


def portfolio_tables(pos_cliente: pd.DataFrame, p: dict[str, float], pl: float) -> tuple[pd.DataFrame, pd.DataFrame]:
    sub_target = subbucket_targets_from_model(p, pl)
    macro_target = macro_targets_from_sub(sub_target, pl, p)

    actual_macro = pos_cliente.groupby("classe_macro", as_index=False)["valor_mercado"].sum().rename(columns={"classe_macro": "Classe", "valor_mercado": "Valor Atual"})
    macro = actual_macro.merge(macro_target, how="outer", on="Classe").fillna(0)
    macro["Peso Atual"] = np.where(pl > 0, macro["Valor Atual"] / pl, 0)
    macro["Diferença"] = macro["Valor Ideal"] - macro["Valor Atual"]
    macro["Status"] = macro.apply(lambda r: status_por_diff(r["Diferença"], r["Valor Ideal"]), axis=1)
    macro["Ação"] = macro["Diferença"].apply(acao_por_diff)
    macro = macro[["Classe", "Status", "Peso Atual", "Peso Ideal", "Valor Atual", "Valor Ideal", "Diferença", "Ação"]]

    actual_sub = pos_cliente.groupby(["classe_macro", "subbucket"], as_index=False)["valor_mercado"].sum()
    actual_sub = actual_sub.rename(columns={"classe_macro": "Classe", "subbucket": "Subbucket", "valor_mercado": "Valor Atual"})
    sub = actual_sub.merge(sub_target, how="outer", on=["Classe", "Subbucket"]).fillna(0)
    sub["Peso Atual"] = np.where(pl > 0, sub["Valor Atual"] / pl, 0)
    sub["Diferença"] = sub["Valor Ideal"] - sub["Valor Atual"]
    sub["Status"] = sub.apply(lambda r: status_por_diff(r["Diferença"], r["Valor Ideal"]), axis=1)
    sub["Prioridade"] = sub["Diferença"].apply(lambda x: prioridade_por_diff(x, pl))
    sub["Ação"] = sub["Diferença"].apply(acao_por_diff)
    order = {b: i for i, b in enumerate(SUBBUCKET_ORDER)}
    sub["_ord"] = sub["Subbucket"].map(order).fillna(999)
    sub = sub.sort_values(["Classe", "_ord", "Subbucket"]).drop(columns="_ord")
    sub = sub[["Classe", "Subbucket", "Status", "Prioridade", "Peso Atual", "Peso Ideal", "Valor Atual", "Valor Ideal", "Diferença", "Ação"]]
    return macro, sub



def rv_recommendation(pos_cliente: pd.DataFrame, p: dict[str, float], pl: float, modelo: str, price_ref: dict[str, float] | None = None) -> pd.DataFrame:
    price_ref = price_ref or {}
    rows = []
    universe_by_bucket = rv_universe(modelo)
    model_tickers = {ticker_clean(t) for xs in universe_by_bucket.values() for t in xs}
    for bucket, tickers in universe_by_bucket.items():
        alvo_total = pl * peso_get(p, bucket)
        alvo_ativo = alvo_total / len(tickers) if tickers else 0
        for t in tickers:
            preco, qtd, atual = current_exchange_position(pos_cliente, t, price_ref)
            can_rebalance = ticker_can_rebalance(pos_cliente, t)
            uses_qty = ticker_uses_quantity(pos_cliente, t)
            qtd_ideal = round(alvo_ativo / preco) if uses_qty and pd.notna(preco) and preco > 0 else np.nan
            qtd_operar = (qtd_ideal - qtd) if can_rebalance and pd.notna(qtd_ideal) else (0.0 if not can_rebalance else np.nan)
            diff = alvo_ativo - atual if pd.notna(atual) else np.nan
            rows.append([t, preco, qtd, atual, qtd_ideal, alvo_ativo, diff, qtd_operar, bucket])

    bolsa_mask = pos_cliente["subbucket"].isin(["Ações", "FIIs"]) & pos_cliente["ticker_norm"].apply(is_valid_b3_ticker)
    nonlisted_mask = pos_cliente.get("b3_tipo_produto", pd.Series("", index=pos_cliente.index)).astype(str).map(norm).str.contains("CETIPADO|NAO LISTADO|NÃO LISTADO", regex=True, na=False)
    bolsa = pos_cliente[bolsa_mask & ~nonlisted_mask].copy()
    for tk, grp in bolsa.groupby("ticker_norm"):
        if not tk or tk in model_tickers:
            continue
        ativo = str(grp["asset_id"].iloc[0])
        preco, qtd, atual = current_exchange_position(pos_cliente, tk, price_ref)
        can_rebalance = ticker_can_rebalance(pos_cliente, tk)
        uses_qty = ticker_uses_quantity(pos_cliente, tk)
        diff = -atual if pd.notna(atual) else np.nan
        qtd_operar = (-qtd if uses_qty and pd.notna(preco) and preco > 0 else np.nan) if can_rebalance else 0.0
        rows.append([ativo, preco, qtd, atual, 0, 0.0, diff, qtd_operar, "Fora do modelo"])
    return pd.DataFrame(rows, columns=["Ativo", "Preço referência", "Qtd Atual", "Valor Atual", "Qtd Ideal", "Valor Ideal", "Diferença", "Qtd a operar", "Grupo"])


def fiinfra_recommendation(pos_cliente: pd.DataFrame, p: dict[str, float], pl: float, price_ref: dict[str, float] | None = None) -> pd.DataFrame:
    price_ref = price_ref or {}
    rows = []
    model_tickers = {ticker_clean(t) for t in FI_INFRA_TICKERS}
    grupos = [
        ("Infraestrutura pós-fixada", peso_get(p, "FiInfra e Cetipados"), FI_INFRA_POS_TICKERS),
        ("Infraestrutura indexada à inflação", peso_get(p, "FiInfra e Cetipado"), FI_INFRA_INFLACAO_TICKERS),
    ]
    for nome, peso_total, tickers in grupos:
        alvo_total = pl * peso_total
        if alvo_total <= 0 or not tickers:
            continue
        alvo_ativo = alvo_total / len(tickers)
        for t in tickers:
            preco, qtd, atual = current_exchange_position(pos_cliente, t, price_ref)
            can_rebalance = ticker_can_rebalance(pos_cliente, t)
            uses_qty = ticker_uses_quantity(pos_cliente, t)
            qtd_ideal = round(alvo_ativo / preco) if uses_qty and pd.notna(preco) and preco > 0 else np.nan
            qtd_operar = (qtd_ideal - qtd) if can_rebalance and pd.notna(qtd_ideal) else (0.0 if not can_rebalance else np.nan)
            diff = alvo_ativo - atual if pd.notna(atual) else np.nan
            rows.append([t, preco, qtd, atual, qtd_ideal, alvo_ativo, diff, qtd_operar, nome])

    held_mask = pos_cliente["subbucket"].isin(["FiInfra Pós", "FiInfra Inflação", "FiInfra e Cetipados"]) & pos_cliente["ticker_norm"].apply(is_valid_b3_ticker)
    nonlisted_mask = pos_cliente.get("b3_tipo_produto", pd.Series("", index=pos_cliente.index)).astype(str).map(norm).str.contains("CETIPADO|NAO LISTADO|NÃO LISTADO", regex=True, na=False)
    held = pos_cliente[held_mask & ~nonlisted_mask].copy()
    for tk, grp in held.groupby("ticker_norm"):
        if not tk or tk in model_tickers:
            continue
        ativo = str(grp["asset_id"].iloc[0])
        preco, qtd, atual = current_exchange_position(pos_cliente, tk, price_ref)
        can_rebalance = ticker_can_rebalance(pos_cliente, tk)
        uses_qty = ticker_uses_quantity(pos_cliente, tk)
        diff = -atual if pd.notna(atual) else np.nan
        qtd_operar = (-qtd if uses_qty and pd.notna(preco) and preco > 0 else np.nan) if can_rebalance else 0.0
        rows.append([ativo, preco, qtd, atual, 0, 0.0, diff, qtd_operar, "Fora do modelo"])
    return pd.DataFrame(rows, columns=["Ativo", "Preço referência", "Qtd Atual", "Valor Atual", "Qtd Ideal", "Valor Ideal", "Diferença", "Qtd a operar", "Grupo"])

def action_summary(pos_cliente: pd.DataFrame, sub_df: pd.DataFrame, pl: float) -> tuple[pd.DataFrame, float, float]:
    saldo = float(pos_cliente.loc[pos_cliente["subbucket"].eq("Saldo em Conta"), "valor_mercado"].sum())
    fora_liquidez = float(pos_cliente.loc[pos_cliente["classe_macro"].eq("Fora da Estratégia"), "valor_mercado"].sum())
    compras = sub_df[(sub_df["Diferença"] > 300) & (~sub_df["Classe"].isin(["Caixa", "Fora da Estratégia"]))].copy()
    vendas = sub_df[(sub_df["Diferença"] < -300) & (~sub_df["Classe"].isin(["Caixa"]))].copy()
    rows = []
    for _, r in vendas.sort_values("Diferença").head(8).iterrows():
        rows.append(["Liberar", r["Subbucket"], r["Classe"], abs(float(r["Diferença"])), "Reduzir excesso"])
    for _, r in compras.sort_values("Diferença", ascending=False).head(8).iterrows():
        rows.append(["Alocar", r["Subbucket"], r["Classe"], float(r["Diferença"]), "Comprar/aportar"])
    acao = pd.DataFrame(rows, columns=["Tipo", "Destino/Origem", "Classe", "Valor", "O que fazer"])
    return acao, saldo, fora_liquidez



PARENT_WEIGHT_KEYS = {
    "RF POS", "RF PÓS", "FUNDOS DE INVEST.", "FUNDOS DE INVEST", "RF PRE", "RF PRÉ",
    "RF INFLACAO", "RF INFLAÇÃO", "RV BRASIL", "INTERNACIONAL"
}

GROUP_DESCRIPTIONS = {
    "Renda fixa pós-fixada": "Parcela voltada para liquidez, estabilidade e acompanhamento das taxas de juros. Em geral, utiliza instrumentos atrelados ao CDI ou à Selic, distribuídos por prazo de resgate.",
    "Renda fixa prefixada": "Parcela com taxa conhecida no momento da aplicação. Pode ajudar a travar juros quando o cenário for favorável, respeitando o prazo e o perfil do cliente.",
    "Renda fixa indexada à inflação": "Parcela que busca preservar poder de compra ao longo do tempo, combinando uma taxa real com a variação da inflação.",
    "Fundos de infraestrutura e crédito incentivado": "Parcela destinada a instrumentos ligados a infraestrutura e crédito incentivado, com potencial de diversificação e geração de renda.",
    "Crédito privado": "Parcela destinada a títulos emitidos por empresas ou estruturas de crédito, buscando retorno adicional mediante análise de risco.",
    "Ações brasileiras": "Parcela de crescimento da carteira, composta por empresas listadas na bolsa brasileira conforme a estratégia definida pela gestão.",
    "Fundos Imobiliários / FIAGROs": "Parcela composta por fundos imobiliários e FIAGROs negociados em bolsa, com potencial de geração de renda e diversificação entre imóveis, recebíveis e agronegócio.",
    "Investimentos internacionais": "Parcela voltada à diversificação geográfica e cambial, reduzindo a dependência exclusiva do mercado brasileiro.",
    "Outros instrumentos": "Parcela complementar para instrumentos específicos, quando aplicável à estratégia selecionada.",
}


def friendly_group_for_weight(key: str) -> tuple[str | None, str | None]:
    k = norm(key)
    if k in PARENT_WEIGHT_KEYS:
        return None, None
    if k == "IMEDIATO":
        return "Renda fixa pós-fixada", "Liquidez imediata"
    if "1 A 30" in k:
        return "Renda fixa pós-fixada", "Resgate entre 1 e 30 dias"
    if "31 A 180" in k:
        return "Renda fixa pós-fixada", "Resgate entre 31 e 180 dias"
    if "181 A 360" in k:
        return "Renda fixa pós-fixada", "Resgate entre 181 e 360 dias"
    if "361" in k:
        return "Renda fixa pós-fixada", "Resgate acima de 361 dias"
    if "BANCARIO PRE" in k or "BANCÁRIO PRÉ" in k or "BANCARIO PRÉ" in k:
        return "Renda fixa prefixada", "Títulos bancários prefixados"
    if "TESOURO PRE" in k or "TESOURO PRÉ" in k:
        return "Renda fixa prefixada", "Tesouro prefixado"
    if k == "BANCARIO" or k == "BANCÁRIO":
        return "Renda fixa indexada à inflação", "Títulos bancários indexados à inflação"
    if k == "TESOURO":
        return "Renda fixa indexada à inflação", "Tesouro indexado à inflação"
    if "FIINFRA" in k or "CETIP" in k:
        # Na planilha existem duas alocações estratégicas diferentes.
        # Mantemos as duas separadas para não duplicar a carteira no PDF/teórica.
        if "CETIPADOS" in k:
            return "Fundos de infraestrutura e crédito incentivado", "Infraestrutura pós-fixada"
        if "CETIPADO" in k:
            return "Fundos de infraestrutura e crédito incentivado", "Infraestrutura indexada à inflação"
        if "POS" in k or "PÓS" in k:
            return "Fundos de infraestrutura e crédito incentivado", "Infraestrutura pós-fixada"
        if "INFL" in k or "IPCA" in k:
            return "Fundos de infraestrutura e crédito incentivado", "Infraestrutura indexada à inflação"
        return "Fundos de infraestrutura e crédito incentivado", "Fundos de infraestrutura e crédito incentivado"
    if "CREDITO PRIVADO" in k or "CRÉDITO PRIVADO" in k:
        return "Crédito privado", "Crédito privado"
    if "ACOES" in k or "AÇÕES" in k or k == "ACOES" or k == "AÇÕES":
        return "Ações brasileiras", "Carteira de ações brasileiras"
    if "FII" in k:
        return "Fundos Imobiliários / FIAGROs", "Fundos Imobiliários / FIAGROs"
    if k == "RENDA FIXA":
        return "Investimentos internacionais", "Renda fixa internacional"
    if k == "RENDA VARIAVEL" or k == "RENDA VARIÁVEL":
        return "Investimentos internacionais", "Renda variável internacional"
    return "Outros instrumentos", str(key).strip()


def component_explanation(group: str, component: str) -> str:
    if group == "Renda fixa pós-fixada":
        return "Componente de liquidez e estabilidade, organizado conforme o prazo de disponibilidade dos recursos."
    if group == "Renda fixa prefixada":
        return "Componente com taxa definida no início da aplicação."
    if group == "Renda fixa indexada à inflação":
        return "Componente voltado à proteção de poder de compra."
    if group == "Fundos de infraestrutura e crédito incentivado":
        if "pós" in component.lower():
            return "Parcela voltada a instrumentos de infraestrutura com comportamento mais próximo ao CDI e foco em geração de renda."
        if "inflação" in component.lower():
            return "Parcela voltada a instrumentos de infraestrutura indexados à inflação, buscando proteção de poder de compra."
        return "Seleção estratégica de instrumentos ligados a infraestrutura e crédito incentivado."
    if group == "Ações brasileiras":
        return "Carteira estratégica de empresas brasileiras definida pela gestão."
    if group == "Fundos Imobiliários / FIAGROs":
        return "Carteira estratégica de fundos imobiliários definida pela gestão."
    if group == "Investimentos internacionais":
        return "Exposição internacional para diversificação geográfica e cambial."
    return "Componente complementar da estratégia."


def display_order_group(group: str) -> int:
    order = [
        "Renda fixa pós-fixada", "Renda fixa prefixada", "Renda fixa indexada à inflação",
        "Fundos de infraestrutura e crédito incentivado", "Crédito privado", "Ações brasileiras",
        "Fundos Imobiliários / FIAGROs", "Investimentos internacionais", "Outros instrumentos"
    ]
    return order.index(group) if group in order else 999


def fiinfra_tickers_for_component(component: str) -> list[str]:
    c = norm(component)
    if "POS" in c or "PÓS" in c:
        return FI_INFRA_POS_TICKERS
    if "INFL" in c or "IPCA" in c:
        return FI_INFRA_INFLACAO_TICKERS
    return FI_INFRA_TICKERS


def theoretical_portfolio(p: dict[str, float], valor: float, modelo: str) -> pd.DataFrame:
    rows = []
    for key, weight in p.items():
        w = float(weight or 0)
        if w <= 0:
            continue
        group, component = friendly_group_for_weight(key)
        if not group or not component:
            continue
        rows.append({
            "Grupo": group,
            "Composição": component,
            "Nível": "Composição",
            "Ativo": "",
            "Peso": w,
            "Valor": valor * w,
            "Explicação": component_explanation(group, component),
        })

        if group == "Ações brasileiras":
            tickers = rv_universe(modelo).get("Ações", [])
            for t in tickers:
                rows.append({"Grupo": group, "Composição": component, "Nível": "Ativo", "Ativo": t, "Peso": w / len(tickers) if tickers else 0, "Valor": valor * w / len(tickers) if tickers else 0, "Explicação": "Empresa brasileira da carteira estratégica."})
        elif group == "Fundos Imobiliários / FIAGROs":
            tickers = rv_universe(modelo).get("FIIs", [])
            for t in tickers:
                rows.append({"Grupo": group, "Composição": component, "Nível": "Ativo", "Ativo": t, "Peso": w / len(tickers) if tickers else 0, "Valor": valor * w / len(tickers) if tickers else 0, "Explicação": "Fundo imobiliário da carteira estratégica."})
        elif group == "Fundos de infraestrutura e crédito incentivado":
            tickers = fiinfra_tickers_for_component(component)
            for t in tickers:
                rows.append({"Grupo": group, "Composição": component, "Nível": "Ativo", "Ativo": t, "Peso": w / len(tickers) if tickers else 0, "Valor": valor * w / len(tickers) if tickers else 0, "Explicação": "Instrumento selecionado pela estratégia de infraestrutura e crédito incentivado."})

    df = pd.DataFrame(rows, columns=["Grupo", "Composição", "Nível", "Ativo", "Peso", "Valor", "Explicação"])
    if df.empty:
        return df
    df["_ord"] = df["Grupo"].apply(display_order_group)
    df["_nivel_ord"] = df["Nível"].map({"Composição": 0, "Ativo": 1}).fillna(9)
    return df.sort_values(["_ord", "Composição", "_nivel_ord", "Ativo"]).drop(columns=["_ord", "_nivel_ord"]).reset_index(drop=True)


def portfolio_macro_cliente(df_teor: pd.DataFrame) -> pd.DataFrame:
    base = df_teor[df_teor["Nível"].eq("Composição")].copy()
    if base.empty:
        return pd.DataFrame(columns=["Classe de investimento", "Peso sugerido", "Valor sugerido"])
    macro = base.groupby("Grupo", as_index=False).agg(Peso=("Peso", "sum"), Valor=("Valor", "sum"))
    macro["_ord"] = macro["Grupo"].apply(display_order_group)
    macro = macro.sort_values("_ord").drop(columns="_ord")
    return macro.rename(columns={"Grupo": "Classe de investimento", "Peso": "Peso sugerido", "Valor": "Valor sugerido"})


def pdf_paragraph(value, style_name="Cell", font_size=7.1, bold=False, color="#222222", align=0) -> Paragraph:
    """Cria células com quebra de linha real no ReportLab, evitando texto sobreposto no PDF."""
    font = PDF_FONT_BOLD if bold else PDF_FONT_REGULAR
    txt = escape("" if value is None else str(value)).replace("\n", "<br/>")
    return Paragraph(
        txt,
        ParagraphStyle(
            name=style_name,
            fontName=font,
            fontSize=font_size,
            leading=font_size + 2.2,
            textColor=colors.HexColor(color),
            alignment=align,
            wordWrap="CJK",
        ),
    )


def table_for_pdf(rows: list[list[str]], col_widths: list[float], header_bg="#172b4d", font_size=7.1, numeric_cols: list[int] | None = None) -> Table:
    """Tabela segura para PDF: todas as células são Paragraph, com wrap e padding."""
    numeric_cols = numeric_cols or []
    wrapped = []
    for ri, row in enumerate(rows):
        new_row = []
        for ci, cell in enumerate(row):
            is_header = ri == 0
            align = 1 if is_header else (2 if ci in numeric_cols else 0)
            new_row.append(
                pdf_paragraph(
                    cell,
                    style_name=f"PDFTable_{ri}_{ci}",
                    font_size=font_size if not is_header else max(font_size, 7.2),
                    bold=is_header,
                    color="#FFFFFF" if is_header else "#222222",
                    align=align,
                )
            )
        wrapped.append(new_row)

    tbl = Table(wrapped, colWidths=[w * cm for w in col_widths], repeatRows=1, splitByRow=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor(header_bg)),
        ("GRID", (0,0), (-1,-1), .2, colors.HexColor("#d9dde5")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f7f8fb")]),
    ]))
    return tbl


def build_pdf_teorico(df_teor: pd.DataFrame, modelo: str, valor: float, cliente: str = "") -> BytesIO:
    if not HAS_REPORTLAB:
        raise RuntimeError("ReportLab não está instalado.")
    register_pdf_fonts()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.25 * cm, leftMargin=1.25 * cm, topMargin=1.55 * cm, bottomMargin=1.35 * cm)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="MWCover", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=22, leading=26, textColor=colors.white))
    styles.add(ParagraphStyle(name="MWCoverSub", parent=styles["Normal"], fontName=PDF_FONT_REGULAR, fontSize=10, leading=14, textColor=colors.HexColor("#dce5f8")))
    styles.add(ParagraphStyle(name="MWSection", parent=styles["Heading2"], fontName=PDF_FONT_BOLD, fontSize=12, leading=15, textColor=colors.HexColor("#172b4d"), spaceBefore=10, spaceAfter=5))
    styles.add(ParagraphStyle(name="MWText", parent=styles["Normal"], fontName=PDF_FONT_REGULAR, fontSize=8.5, leading=11, textColor=colors.HexColor("#333333")))
    styles.add(ParagraphStyle(name="MWSmall", parent=styles["Normal"], fontName=PDF_FONT_REGULAR, fontSize=7.2, leading=9.2, textColor=colors.HexColor("#5b6472")))

    def footer(canvas, doc_obj):
        canvas.saveState()
        width, height = A4
        canvas.setStrokeColor(colors.HexColor("#dfe4ed"))
        canvas.line(1.25 * cm, 0.95 * cm, width - 1.25 * cm, 0.95 * cm)
        canvas.setFont(PDF_FONT_REGULAR, 7)
        canvas.setFillColor(colors.HexColor("#667085"))
        canvas.drawString(1.25 * cm, 0.58 * cm, "M Wealth | Estudo de Alocação")
        canvas.drawRightString(width - 1.25 * cm, 0.58 * cm, f"Página {doc_obj.page}")
        canvas.restoreState()

    story = []
    lp = logo_pdf_path()
    logo_flow = Image(str(lp), width=4.7 * cm, height=1.45 * cm, kind="proportional") if lp else Spacer(1, .1 * cm)
    cover_left = [
        logo_flow,
        Spacer(1, .35 * cm),
        Paragraph("Estudo de Alocação", styles["MWCover"]),
        Paragraph("Carteira teórica personalizada e organizada por objetivos, liquidez e classes de investimento.", styles["MWCoverSub"]),
        Spacer(1, .5 * cm),
        Paragraph(f"<b>Cliente:</b> {escape(cliente.strip() or 'Não informado')}<br/><b>Perfil:</b> {escape(modelo)}<br/><b>Valor analisado:</b> {escape(format_brl(valor))}<br/><b>Data:</b> {datetime.now().strftime('%d/%m/%Y')}", styles["MWCoverSub"]),
    ]
    banner = Table([[cover_left]], colWidths=[17.1 * cm], rowHeights=[6.7 * cm])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#172b4d")),
        ("BOX", (0,0), (-1,-1), 0, colors.HexColor("#172b4d")),
        ("LEFTPADDING", (0,0), (-1,-1), 20), ("RIGHTPADDING", (0,0), (-1,-1), 20),
        ("TOPPADDING", (0,0), (-1,-1), 18), ("BOTTOMPADDING", (0,0), (-1,-1), 18),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(banner)
    story.append(Spacer(1, .45 * cm))

    macro = portfolio_macro_cliente(df_teor)
    cards = []
    for _, r in macro.head(4).iterrows():
        cards.append([pdf_paragraph(r["Classe de investimento"], font_size=7.2, bold=True, color="#172b4d"), pdf_paragraph(fmt_pct(r["Peso sugerido"]), font_size=11, bold=True, color="#172b4d", align=1), pdf_paragraph(format_brl(r["Valor sugerido"]), font_size=7.5, color="#475467", align=1)])
    if cards:
        card_tbl = Table(cards, colWidths=[7.2 * cm, 3.1 * cm, 5.0 * cm])
        card_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#f4f6fa")),
            ("GRID", (0,0), (-1,-1), .25, colors.HexColor("#dfe4ed")),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING", (0,0), (-1,-1), 8), ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ("TOPPADDING", (0,0), (-1,-1), 7), ("BOTTOMPADDING", (0,0), (-1,-1), 7),
        ]))
        story.append(card_tbl)

    story.append(Paragraph("Resumo executivo", styles["MWSection"]))
    story.append(Paragraph(
        "A carteira apresentada traduz o perfil selecionado em uma distribuição objetiva entre liquidez, proteção, geração de renda, crescimento e diversificação internacional. Os valores são indicativos e devem ser confirmados conforme disponibilidade dos produtos, condições de mercado e particularidades do investidor.",
        styles["MWText"],
    ))

    hierarchy = theoretical_hierarchy_table({r["Composição"]: r["Peso"] for _, r in df_teor[df_teor["Nível"].eq("Composição")].iterrows()}, valor)
    # Usa diretamente a estrutura do dataframe teórico quando as chaves amigáveis não são as do modelo.
    data = [["Estratégia", "Peso sugerido", "Valor sugerido"]]
    for _, r in macro.iterrows():
        data.append([r["Classe de investimento"], fmt_pct(r["Peso sugerido"]), format_brl(r["Valor sugerido"])])
    story.append(Paragraph("Alocação recomendada", styles["MWSection"]))
    story.append(table_for_pdf(data, [8.4, 3.0, 4.2], font_size=7.8, numeric_cols=[1,2]))
    story.append(PageBreak())

    for group in macro["Classe de investimento"].tolist():
        comp = df_teor[(df_teor["Grupo"].eq(group)) & (df_teor["Nível"].eq("Composição"))].copy()
        ativos = df_teor[(df_teor["Grupo"].eq(group)) & (df_teor["Nível"].eq("Ativo"))].copy()
        story.append(Paragraph(group, styles["MWSection"]))
        story.append(Paragraph(GROUP_DESCRIPTIONS.get(group, "Componente da estratégia de alocação."), styles["MWText"]))
        story.append(Spacer(1, .12 * cm))
        data = [["Estratégia", "Peso", "Valor", "Objetivo"]]
        for _, r in comp.iterrows():
            data.append([r["Composição"], fmt_pct(r["Peso"]), format_brl(r["Valor"]), r["Explicação"]])
        story.append(table_for_pdf(data, [4.5, 2.0, 3.0, 7.2], font_size=6.9, numeric_cols=[1,2]))
        if not ativos.empty:
            story.append(Spacer(1, .16 * cm))
            data = [["Produto / ativo", "Estratégia", "Peso", "Valor"]]
            for _, r in ativos.iterrows():
                data.append([r["Ativo"], r["Composição"], fmt_pct(r["Peso"]), format_brl(r["Valor"])])
            story.append(table_for_pdf(data, [3.2, 6.1, 2.2, 3.5], font_size=7.0, numeric_cols=[2,3]))
        story.append(Spacer(1, .25 * cm))

    story.append(Paragraph("Premissas e observações", styles["MWSection"]))
    story.append(Paragraph("A indicação de produtos depende do cadastro mestre, da corretora disponível, dos limites de concentração, dos aportes mínimos e das restrições específicas do cliente. Produtos podem ser substituídos por equivalentes do mesmo objetivo sem alterar a arquitetura estratégica.", styles["MWText"]))
    story.append(Paragraph("Disclaimer", styles["MWSection"]))
    story.append(Paragraph("Este material é meramente informativo e não constitui promessa de rentabilidade. A composição final depende da análise individual do investidor, suitability, disponibilidade de produtos e condições de mercado. Rentabilidade passada não representa garantia de rentabilidade futura.", styles["MWSmall"]))
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    buf.seek(0)
    return buf

def mapping_files_signature() -> tuple:
    """Invalida o cache quando o arquivo escolhido ou seu conteúdo mudar."""
    fp = master_products_path()
    mp = find_file("Manual de Alocação.xlsx")
    return (
        str(fp.resolve()) if fp.exists() else str(fp),
        *file_cache_signature(fp),
        str(mp.resolve()) if mp.exists() else str(mp),
        *file_cache_signature(mp),
    )




# =============================================================================
# Pool de produtos, restrições e governança de modelos
# =============================================================================
@st.cache_data(show_spinner=False)
def load_product_pool_cached(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> pd.DataFrame:
    path = Path(path_str)
    if not path.exists():
        return pd.DataFrame()
    try:
        xls = pd.ExcelFile(path)
        if "Pool de Produtos" not in xls.sheet_names:
            return pd.DataFrame()
        df = pd.read_excel(path, sheet_name="Pool de Produtos")
        df.columns = [str(c).strip() for c in df.columns]
        expected = [
            "ID_PRODUTO", "NOME_PRODUTO", "TIPO_IDENTIFICADOR", "IDENTIFICADOR",
            "CLASSE", "SUBBUCKET", "POOL", "CORRETORA", "PERFIL_MINIMO",
            "PRIORIDADE_COMPRA", "PESO_NO_POOL", "APORTE_MINIMO",
            "LIMITE_POR_CLIENTE", "ELEGIVEL_COMPRA", "APENAS_MANUTENCAO",
            "PRODUTO_FECHADO", "INDEXADOR", "GESTORA_EMISSOR", "SETOR", "OBSERVACAO",
        ]
        df = canonicalize_master_columns(df, expected)
        for c in expected:
            if c not in df.columns:
                df[c] = np.nan if c in {"PRIORIDADE_COMPRA", "PESO_NO_POOL", "APORTE_MINIMO", "LIMITE_POR_CLIENTE"} else ""
        df = df[df["NOME_PRODUTO"].fillna("").astype(str).str.strip().ne("")].copy()
        df["IDENTIFICADOR_NORM"] = df["IDENTIFICADOR"].fillna("").astype(str).map(norm)
        df["PRIORIDADE_COMPRA"] = pd.to_numeric(df["PRIORIDADE_COMPRA"], errors="coerce").fillna(999)
        df["PESO_NO_POOL"] = pd.to_numeric(df["PESO_NO_POOL"], errors="coerce").fillna(0.0)
        df["APORTE_MINIMO"] = pd.to_numeric(df["APORTE_MINIMO"], errors="coerce").fillna(0.0)
        df["LIMITE_POR_CLIENTE"] = pd.to_numeric(df["LIMITE_POR_CLIENTE"], errors="coerce").fillna(1.0)
        return df.reset_index(drop=True)
    except Exception as exc:
        st.session_state["pool_load_error"] = str(exc)
        return pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_client_restrictions_cached(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> pd.DataFrame:
    path = Path(path_str)
    if not path.exists():
        return pd.DataFrame()
    try:
        xls = pd.ExcelFile(path)
        if "Restrições por Cliente" not in xls.sheet_names:
            return pd.DataFrame()
        df = pd.read_excel(path, sheet_name="Restrições por Cliente")
        df.columns = [str(c).strip() for c in df.columns]
        expected = ["GRUPO_CLIENTE", "NIVEL", "TIPO_REGRA", "IDENTIFICADOR", "ACAO", "LIMITE_PERCENTUAL", "SUBSTITUTO", "ATIVO", "DATA_INICIO", "DATA_FIM", "MOTIVO", "STATUS"]
        df = canonicalize_master_columns(df, expected)
        for c in expected:
            if c not in df.columns:
                df[c] = ""
        df = df[df["GRUPO_CLIENTE"].fillna("").astype(str).str.strip().ne("")].copy()
        df["LIMITE_PERCENTUAL"] = pd.to_numeric(df["LIMITE_PERCENTUAL"], errors="coerce")
        return df.reset_index(drop=True)
    except Exception as exc:
        st.session_state["restriction_load_error"] = str(exc)
        return pd.DataFrame()


def applicable_restrictions(restrictions: pd.DataFrame, grupo: str, cliente: str = "", conta: str = "") -> pd.DataFrame:
    if restrictions.empty:
        return restrictions.copy()
    now = pd.Timestamp.now().normalize()
    out = restrictions.copy()
    out = out[~out["STATUS"].fillna("").astype(str).map(norm).isin({"INATIVA", "EXEMPLO"})]
    target = out["GRUPO_CLIENTE"].fillna("").astype(str).map(norm)
    level = out["NIVEL"].fillna("").astype(str).map(norm)
    match = (
        (level.eq("GRUPO") & target.eq(norm(grupo))) |
        (level.eq("CLIENTE") & target.eq(norm(cliente))) |
        (level.eq("CONTA") & target.eq(norm(conta)))
    )
    out = out[match].copy()
    if "DATA_INICIO" in out.columns:
        ini = pd.to_datetime(out["DATA_INICIO"], errors="coerce")
        out = out[ini.isna() | ini.le(now)]
    if "DATA_FIM" in out.columns:
        fim = pd.to_datetime(out["DATA_FIM"], errors="coerce")
        out = out[fim.isna() | fim.ge(now)]
    return out



POOL_COLUMNS = [
    "ID_PRODUTO", "NOME_PRODUTO", "TIPO_IDENTIFICADOR", "IDENTIFICADOR",
    "CLASSE", "SUBBUCKET", "POOL", "CORRETORA", "PERFIL_MINIMO",
    "PRIORIDADE_COMPRA", "PESO_NO_POOL", "APORTE_MINIMO",
    "LIMITE_POR_CLIENTE", "ELEGIVEL_COMPRA", "APENAS_MANUTENCAO",
    "PRODUTO_FECHADO", "INDEXADOR", "GESTORA_EMISSOR", "SETOR", "OBSERVACAO",
]
RESTRICTION_COLUMNS = [
    "GRUPO_CLIENTE", "NIVEL", "TIPO_REGRA", "IDENTIFICADOR", "ACAO",
    "LIMITE_PERCENTUAL", "SUBSTITUTO", "ATIVO", "DATA_INICIO", "DATA_FIM",
    "MOTIVO", "STATUS",
]
B3_EDITOR_COLUMNS = [
    "TICKER", "NOME_ATIVO", "TIPO_PRODUTO", "CLASSE_OPERACIONAL",
    "SUBBUCKET_OPERACIONAL", "ESTRATEGIA", "REBALANCEAR",
    "STATUS_MAPEAMENTO", "OBSERVACAO_OPERACIONAL", "LIQUIDEZ_OPERACIONAL",
    "FONTE_PRECO", "SETOR",
]


def read_master_sheet_for_editor(sheet_name: str, expected: list[str]) -> pd.DataFrame:
    path = master_products_path()
    if not path.exists():
        return pd.DataFrame(columns=expected)
    try:
        raw = pd.read_excel(path, sheet_name=sheet_name)
        raw.columns = [str(c).strip() for c in raw.columns]
        raw = canonicalize_master_columns(raw, expected)
        for c in expected:
            if c not in raw.columns:
                raw[c] = np.nan if c in {"PRIORIDADE_COMPRA", "PESO_NO_POOL", "APORTE_MINIMO", "LIMITE_POR_CLIENTE", "LIMITE_PERCENTUAL", "LIQUIDEZ_OPERACIONAL"} else ""
        return raw[expected].copy()
    except Exception:
        return pd.DataFrame(columns=expected)


def _clean_editor_df(df: pd.DataFrame, required_col: str, columns: list[str]) -> pd.DataFrame:
    out = df.copy()
    for c in columns:
        if c not in out.columns:
            out[c] = ""
    out = out[columns].copy()
    out = out[out[required_col].fillna("").astype(str).str.strip().ne("")].copy()
    return out.reset_index(drop=True)


def save_master_sheet_from_app(sheet_name: str, df: pd.DataFrame, columns: list[str], required_col: str) -> tuple[bool, str]:
    """Salva uma aba operacional preservando o restante do Cadastro Mestre."""
    path = master_products_path()
    if not path.exists():
        return False, f"Cadastro Mestre não encontrado: {path}"
    try:
        from openpyxl import load_workbook
        from copy import copy as _copy
        clean = _clean_editor_df(df, required_col, columns)
        wb = load_workbook(path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]
        # Preserva o estilo da primeira linha e, quando possível, da primeira linha de dados.
        for j, col in enumerate(columns, start=1):
            ws.cell(1, j).value = col
        max_clear_row = max(ws.max_row, len(clean) + 5)
        max_clear_col = max(ws.max_column, len(columns))
        for row in ws.iter_rows(min_row=2, max_row=max_clear_row, min_col=1, max_col=max_clear_col):
            for cell in row:
                cell.value = None
        template_row = 2 if ws.max_row >= 2 else None
        for i, record in enumerate(clean.itertuples(index=False, name=None), start=2):
            for j, value in enumerate(record, start=1):
                cell = ws.cell(i, j)
                if isinstance(value, float) and pd.isna(value):
                    value = None
                cell.value = value
                if template_row and i > 2:
                    src = ws.cell(template_row, j)
                    if src.has_style:
                        cell._style = _copy(src._style)
                        if src.number_format:
                            cell.number_format = src.number_format
        tmp = path.with_name(path.stem + "__tmp_app" + path.suffix)
        wb.save(tmp)
        tmp.replace(path)
        st.cache_data.clear()
        return True, f"{sheet_name}: {len(clean)} registro(s) salvos no Cadastro Mestre."
    except PermissionError:
        return False, "Não foi possível salvar. Feche o Cadastro Mestre no Excel e tente novamente."
    except Exception as exc:
        return False, f"Falha ao salvar {sheet_name}: {exc}"


def update_fund_override_from_app(sheet_name: str, row_index: int, updates: dict[str, object]) -> tuple[bool, str]:
    path = master_products_path()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb[sheet_name]
        headers = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
        excel_row = int(row_index) + 2
        for key, value in updates.items():
            col = headers.get(norm(key))
            if col is None:
                col = ws.max_column + 1
                ws.cell(1, col).value = key
                headers[norm(key)] = col
            ws.cell(excel_row, col).value = None if (isinstance(value, float) and pd.isna(value)) else value
        tmp = path.with_name(path.stem + "__tmp_app" + path.suffix)
        wb.save(tmp)
        tmp.replace(path)
        st.cache_data.clear()
        return True, "Override salvo no Cadastro Mestre."
    except PermissionError:
        return False, "Feche o Cadastro Mestre no Excel antes de salvar pelo app."
    except Exception as exc:
        return False, f"Falha ao salvar override: {exc}"


def current_value_for_pool_product(pos_cliente: pd.DataFrame, prod: pd.Series) -> float:
    """Calcula a posição atual do produto do pool usando a melhor chave disponível.

    Para fundos por CNPJ, primeiro usa o CNPJ bruto da corretora; se ele não
    existir, usa o CNPJ do casamento com o Cadastro Mestre e, por último, o nome
    normalizado já reconciliado. Isso evita recomendar o alvo cheio para um fundo
    que o cliente já possui, mas cujo relatório veio sem CNPJ.
    """
    if pos_cliente.empty:
        return 0.0

    tipo = norm(prod.get("TIPO_IDENTIFICADOR", ""))
    ident = prod.get("IDENTIFICADOR", "")
    nome_produto = str(prod.get("NOME_PRODUTO", "") or "").strip()
    idx = pos_cliente.index
    mask = pd.Series(False, index=idx)

    if tipo == "CNPJ":
        key = only_digits_str(ident)
        if key:
            raw_cnpj = pos_cliente.get("cnpj", pd.Series("", index=idx)).apply(only_digits_str)
            mapped_cnpj = pos_cliente.get("manual_cnpj", pd.Series("", index=idx)).apply(only_digits_str)
            mask = raw_cnpj.eq(key) | mapped_cnpj.eq(key)

        if not mask.any() and nome_produto:
            key_name = fund_name_key(nome_produto)
            if key_name:
                manual_names = pos_cliente.get("manual_fundo", pd.Series("", index=idx)).astype(str).apply(fund_name_key)
                asset_names = pos_cliente.get("asset_nome", pd.Series("", index=idx)).astype(str).apply(fund_name_key)
                asset_ids = pos_cliente.get("asset_id", pd.Series("", index=idx)).astype(str).apply(fund_name_key)
                mask = manual_names.eq(key_name) | asset_names.eq(key_name) | asset_ids.eq(key_name)

    elif tipo == "NOME":
        key = fund_name_key(str(ident or nome_produto))
        if key:
            manual_names = pos_cliente.get("manual_fundo", pd.Series("", index=idx)).astype(str).apply(fund_name_key)
            asset_names = pos_cliente.get("asset_nome", pd.Series("", index=idx)).astype(str).apply(fund_name_key)
            asset_ids = pos_cliente.get("asset_id", pd.Series("", index=idx)).astype(str).apply(fund_name_key)
            mask = manual_names.eq(key) | asset_names.eq(key) | asset_ids.eq(key)

    else:
        key = ticker_clean(ident)
        series = pos_cliente.get("ticker_norm", pd.Series("", index=idx)).astype(str)
        mask = series.eq(key) if key else pd.Series(False, index=idx)

    if not mask.any():
        return 0.0
    return float(pd.to_numeric(pos_cliente.loc[mask, "valor_mercado"], errors="coerce").fillna(0.0).sum())


def pool_metadata_for_ticker(ticker: str, pool: pd.DataFrame) -> dict[str, str]:
    tk = ticker_clean(ticker)
    if pool is not None and not pool.empty:
        ids = pool.get("IDENTIFICADOR", pd.Series("", index=pool.index)).astype(str).apply(ticker_clean)
        hit = pool[ids.eq(tk)]
        if not hit.empty:
            r = hit.iloc[0]
            return {
                "SETOR": str(r.get("SETOR", "")).strip(),
                "GESTORA_EMISSOR": str(r.get("GESTORA_EMISSOR", "")).strip(),
                "CLASSE": str(r.get("CLASSE", "")).strip(),
                "SUBBUCKET": str(r.get("SUBBUCKET", "")).strip(),
                "NOME_PRODUTO": str(r.get("NOME_PRODUTO", ticker)).strip(),
            }
    # Fallback no Cadastro B3 para permitir restrição por classe/setor mesmo
    # quando o ativo ainda não faz parte do Pool de Produtos.
    try:
        path = master_products_path()
        b3 = load_b3_master_cached(str(path), *file_cache_signature(path))
        hit = b3[b3["ticker_norm"].astype(str).eq(tk)]
        if not hit.empty:
            r = hit.iloc[0]
            return {
                "SETOR": str(r.get("b3_setor", "")).strip(),
                "GESTORA_EMISSOR": "",
                "CLASSE": str(r.get("b3_classe_operacional", "")).strip(),
                "SUBBUCKET": str(r.get("b3_subbucket_operacional", "")).strip(),
                "NOME_PRODUTO": str(r.get("b3_nome", ticker)).strip(),
            }
    except Exception:
        pass
    return {}

def restriction_action_for_product(row: pd.Series, restrictions: pd.DataFrame) -> tuple[str, str, float | None]:
    if restrictions.empty:
        return "", "", None
    candidates = {
        "TICKER": norm(row.get("IDENTIFICADOR", row.get("Ativo", ""))),
        "CNPJ": only_digits_str(row.get("IDENTIFICADOR", "")),
        "PRODUTO": norm(row.get("NOME_PRODUTO", row.get("Ativo", ""))),
        "SETOR": norm(row.get("SETOR", "")),
        "GESTORA": norm(row.get("GESTORA_EMISSOR", "")),
        "SUBBUCKET": norm(row.get("SUBBUCKET", row.get("Grupo", ""))),
        "CLASSE": norm(row.get("CLASSE", row.get("Classe", ""))),
    }
    for _, rr in restrictions.iterrows():
        tipo = norm(rr.get("TIPO_REGRA", ""))
        ident = norm(rr.get("IDENTIFICADOR", ""))
        if tipo == "CNPJ":
            matched = only_digits_str(rr.get("IDENTIFICADOR", "")) == candidates["CNPJ"] and bool(candidates["CNPJ"])
        else:
            matched = ident and ident == candidates.get(tipo, "")
        if matched:
            limite = pd.to_numeric(rr.get("LIMITE_PERCENTUAL", np.nan), errors="coerce")
            return str(rr.get("ACAO", "")).strip(), str(rr.get("MOTIVO", "")).strip(), (float(limite) if pd.notna(limite) else None)
    return "", "", None


def apply_restrictions_to_market_orders(
    df_orders: pd.DataFrame,
    restrictions: pd.DataFrame,
    pool: pd.DataFrame | None = None,
    pl_base: float = 0.0,
) -> pd.DataFrame:
    if df_orders.empty or restrictions.empty:
        return df_orders
    out = df_orders.copy()
    out["Restrição"] = ""
    out["Motivo da restrição"] = ""
    out["Substituto"] = ""
    for idx, row in out.iterrows():
        meta = pool_metadata_for_ticker(str(row.get("Ativo", "")), pool if pool is not None else pd.DataFrame())
        group = str(row.get("Grupo", ""))
        classe_inferida = meta.get("CLASSE", "")
        if not classe_inferida:
            classe_inferida = "RV Brasil" if norm(group) in {"ACOES", "FIIS", "FORA DO MODELO"} else ("RF Brasil" if "INFRA" in norm(group) else "")
        probe = pd.Series({
            "IDENTIFICADOR": row.get("Ativo", ""),
            "NOME_PRODUTO": meta.get("NOME_PRODUTO", row.get("Ativo", "")),
            "SUBBUCKET": meta.get("SUBBUCKET", group),
            "CLASSE": classe_inferida,
            "SETOR": meta.get("SETOR", ""),
            "GESTORA_EMISSOR": meta.get("GESTORA_EMISSOR", ""),
        })
        action, reason, limit_pct = restriction_action_for_product(probe, restrictions)
        action_n = norm(action)
        qtd_op = pd.to_numeric(row.get("Qtd a operar", 0), errors="coerce")
        qtd_atual = pd.to_numeric(row.get("Qtd Atual", 0), errors="coerce")
        valor_atual = pd.to_numeric(row.get("Valor Atual", 0), errors="coerce")
        preco = pd.to_numeric(row.get("Preço referência", np.nan), errors="coerce")

        if action_n == "NAO COMPRAR" and pd.notna(qtd_op) and qtd_op > 0:
            out.at[idx, "Qtd a operar"] = 0
            out.at[idx, "Diferença"] = 0.0
        elif action_n == "NAO VENDER" and pd.notna(qtd_op) and qtd_op < 0:
            out.at[idx, "Qtd a operar"] = 0
            out.at[idx, "Diferença"] = 0.0
        elif action_n == "MANTER POSICAO":
            out.at[idx, "Qtd a operar"] = 0
            out.at[idx, "Qtd Ideal"] = qtd_atual if pd.notna(qtd_atual) else out.at[idx, "Qtd Ideal"]
            out.at[idx, "Valor Ideal"] = valor_atual if pd.notna(valor_atual) else out.at[idx, "Valor Ideal"]
            out.at[idx, "Diferença"] = 0.0
        elif action_n == "EXCLUIR DA CARTEIRA":
            out.at[idx, "Qtd Ideal"] = 0
            out.at[idx, "Valor Ideal"] = 0.0
            out.at[idx, "Diferença"] = -float(valor_atual or 0.0)
            out.at[idx, "Qtd a operar"] = -float(qtd_atual or 0.0) if pd.notna(qtd_atual) else np.nan
        elif action_n == "LIMITAR PERCENTUAL" and limit_pct is not None and pl_base > 0:
            limite_valor = pl_base * float(limit_pct)
            atual = float(valor_atual or 0.0)
            ideal_original = float(pd.to_numeric(row.get("Valor Ideal", 0), errors="coerce") or 0.0)
            ideal_novo = min(ideal_original, limite_valor)
            out.at[idx, "Valor Ideal"] = ideal_novo
            out.at[idx, "Diferença"] = ideal_novo - atual
            if pd.notna(preco) and preco > 0:
                qtd_ideal = round(ideal_novo / float(preco))
                out.at[idx, "Qtd Ideal"] = qtd_ideal
                out.at[idx, "Qtd a operar"] = qtd_ideal - float(qtd_atual or 0.0)
        elif action_n == "SUBSTITUIR POR PRODUTO":
            out.at[idx, "Qtd a operar"] = 0
            out.at[idx, "Diferença"] = 0.0
            matched = restrictions[(restrictions["TIPO_REGRA"].fillna("").astype(str).map(norm).eq("TICKER")) & (restrictions["IDENTIFICADOR"].fillna("").astype(str).map(norm).eq(norm(row.get("Ativo", ""))))]
            if not matched.empty:
                out.at[idx, "Substituto"] = str(matched.iloc[0].get("SUBSTITUTO", "")).strip()
        if action:
            out.at[idx, "Restrição"] = action
            out.at[idx, "Motivo da restrição"] = reason
    return out


def recommend_exact_products(
    sub_df: pd.DataFrame,
    pos_cliente: pd.DataFrame,
    pool: pd.DataFrame,
    restrictions: pd.DataFrame,
    pl_base: float,
    corretoras_cliente: set[str],
) -> pd.DataFrame:
    """Recomenda no máximo um produto do pool por classe/subbucket que precisa de ajuste.

    Lógica em cascata por prioridade (não mais rateio proporcional do valor todo
    entre todos os produtos elegíveis de uma vez):

    - Compra: percorre as prioridades em ordem crescente (1, 2, 3...). O valor
      ideal de cada prioridade é o peso dela (PESO_NO_POOL) aplicado sobre o
      Valor Ideal da classe inteira. Assim que encontra a primeira prioridade
      cujo valor já aplicado está abaixo do ideal dela, recomenda comprar só
      essa - as prioridades seguintes só aparecem quando essa estiver completa.
    - Venda: percorre as prioridades em ordem decrescente (a mais baixa
      primeiro). Assim que encontra a primeira prioridade com valor aplicado
      acima do ideal dela, recomenda vender o excesso só dessa.

    Resultado: uma linha por classe que precisa de ajuste, e não uma lista
    fragmentada em N produtos. PERFIL_MINIMO e APORTE_MINIMO não participam
    da decisão: as restrições de perfil já são definidas pelas carteiras.
    """
    columns = ["Estratégia", "Ação", "Produto recomendado", "Identificador", "Corretora", "Valor atual no produto", "Alvo do produto", "Valor recomendado", "Prioridade", "Peso no pool", "Limite", "Restrição", "Observação"]
    if pool.empty or sub_df.empty or pl_base <= 0:
        return pd.DataFrame(columns=columns)
    rows = []
    needs = sub_df[(pd.to_numeric(sub_df["Diferença"], errors="coerce").abs() > 300) & (~sub_df["Classe"].isin(["Caixa", "Fora da Estratégia"]))]
    for _, need in needs.iterrows():
        candidates = pool[
            pool["CLASSE"].fillna("").astype(str).map(norm).eq(norm(need["Classe"])) &
            pool["SUBBUCKET"].fillna("").astype(str).map(norm).eq(norm(need["Subbucket"]))
        ].copy()
        if candidates.empty:
            continue
        if corretoras_cliente:
            candidates = candidates[candidates["CORRETORA"].fillna("Todas").astype(str).map(norm).isin({"TODAS", "", *{norm(x) for x in corretoras_cliente}})]
        if candidates.empty:
            continue

        allowed = []
        for _, prod in candidates.iterrows():
            action, reason, limit_override = restriction_action_for_product(prod, restrictions)
            if norm(action) == "EXCLUIR DA CARTEIRA":
                continue
            prod = prod.copy()
            prod["_restriction"] = action
            prod["_reason"] = reason
            prod["_limit_override"] = limit_override
            allowed.append(prod)
        if not allowed:
            continue
        candidates = pd.DataFrame(allowed)

        weights = pd.to_numeric(candidates["PESO_NO_POOL"], errors="coerce").clip(lower=0)
        weights = weights.fillna(0)
        if weights.sum() <= 0:
            weights = pd.Series(1.0, index=candidates.index)
        weights = weights / weights.sum()
        candidates = candidates.assign(_peso=weights)

        valor_ideal_classe = float(pd.to_numeric(need.get("Valor Ideal", 0), errors="coerce") or 0.0)
        diff = float(need["Diferença"])

        if diff > 300:
            # Compra: sobe da prioridade 1 em diante.
            compraveis = candidates[
                candidates["ELEGIVEL_COMPRA"].map(lambda x: parse_yes_no(x, False)) &
                ~candidates["APENAS_MANUTENCAO"].map(lambda x: parse_yes_no(x, False)) &
                ~candidates["PRODUTO_FECHADO"].map(lambda x: parse_yes_no(x, False)) &
                ~candidates["_restriction"].map(lambda a: norm(a) == "NAO COMPRAR")
            ].sort_values("PRIORIDADE_COMPRA", ascending=True)
            for _, prod in compraveis.iterrows():
                ideal_tier = valor_ideal_classe * float(prod["_peso"])
                current = current_value_for_pool_product(pos_cliente, prod)
                gap = ideal_tier - current
                if gap <= 50:
                    continue  # esta prioridade já está completa, olha a próxima
                limit_pct = prod.get("_limit_override")
                if limit_pct is None:
                    limit_pct = float(pd.to_numeric(prod.get("LIMITE_POR_CLIENTE", 1), errors="coerce") or 1)
                capacity = max(0.0, pl_base * float(limit_pct) - current)
                amount = min(gap, diff, capacity)
                if amount <= 0:
                    continue
                rows.append([
                    friendly_strategy_name(str(need["Subbucket"])), "Comprar", str(prod["NOME_PRODUTO"]), str(prod["IDENTIFICADOR"]),
                    str(prod["CORRETORA"]), float(current), float(ideal_tier), float(amount), int(prod["PRIORIDADE_COMPRA"]), float(prod["_peso"]),
                    float(limit_pct), str(prod.get("_restriction", "")), str(prod.get("OBSERVACAO", "")),
                ])
                break  # só uma linha por classe

        elif diff < -300:
            # Venda: desce da prioridade mais baixa (número maior) em diante.
            vendiveis = candidates[
                ~candidates["_restriction"].map(lambda a: norm(a) in {"NAO VENDER", "MANTER POSICAO"})
            ].sort_values("PRIORIDADE_COMPRA", ascending=False)
            excesso_total = abs(diff)
            for _, prod in vendiveis.iterrows():
                ideal_tier = valor_ideal_classe * float(prod["_peso"])
                current = current_value_for_pool_product(pos_cliente, prod)
                excesso = current - ideal_tier
                if excesso <= 50:
                    continue  # esta prioridade já está dentro do ideal, olha a próxima
                amount = min(excesso, excesso_total)
                if amount <= 0:
                    continue
                rows.append([
                    friendly_strategy_name(str(need["Subbucket"])), "Vender", str(prod["NOME_PRODUTO"]), str(prod["IDENTIFICADOR"]),
                    str(prod["CORRETORA"]), float(current), float(ideal_tier), -float(amount), int(prod["PRIORIDADE_COMPRA"]), float(prod["_peso"]),
                    np.nan, str(prod.get("_restriction", "")), str(prod.get("OBSERVACAO", "")),
                ])
                break  # só uma linha por classe

    return pd.DataFrame(rows, columns=columns)



@st.cache_data(show_spinner=False)
def load_models_from_master_sheet(path_str: str, mtime_ns: int = 0, file_size: int = 0) -> dict[str, dict[str, float]]:
    path = Path(path_str)
    if not path.exists():
        return {}
    try:
        df = pd.read_excel(path, sheet_name="Modelos de Alocação")
        df.columns = [str(c).strip() for c in df.columns]
        required = {"MODELO", "VERSAO", "STATUS", "SUBBUCKET", "PESO"}
        if not required.issubset(set(df.columns)):
            return {}
        df = df[df["STATUS"].fillna("").astype(str).map(norm).eq("PUBLICADO")].copy()
        if df.empty:
            return {}
        df["VERSAO"] = pd.to_numeric(df["VERSAO"], errors="coerce").fillna(0).astype(int)
        df["PESO"] = pd.to_numeric(df["PESO"], errors="coerce").fillna(0.0)
        result = {}
        for model, grp in df.groupby("MODELO"):
            version = int(grp["VERSAO"].max())
            latest = grp[grp["VERSAO"].eq(version)]
            result[str(model)] = {str(r["SUBBUCKET"]): float(r["PESO"]) for _, r in latest.iterrows() if str(r["SUBBUCKET"]).strip()}
        return result
    except Exception:
        return {}


def sync_published_model_to_master(name: str, weights: dict[str, float], version: int, author: str, reason: str) -> tuple[bool, str]:
    path = master_products_path()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        sheet = "Modelos de Alocação"
        if sheet not in wb.sheetnames:
            ws = wb.create_sheet(sheet)
            ws.append(["MODELO", "VERSAO", "STATUS", "SUBBUCKET", "PESO", "PUBLICADO_EM", "PUBLICADO_POR", "MOTIVO", "OBSERVACAO"])
        ws = wb[sheet]
        headers = {str(ws.cell(1, c).value or "").strip(): c for c in range(1, ws.max_column + 1)}
        required = ["MODELO", "VERSAO", "STATUS", "SUBBUCKET", "PESO", "PUBLICADO_EM", "PUBLICADO_POR", "MOTIVO", "OBSERVACAO"]
        for col in required:
            if col not in headers:
                c = ws.max_column + 1
                ws.cell(1, c).value = col
                headers[col] = c
        # Arquiva versões anteriormente publicadas deste modelo.
        for r in range(2, ws.max_row + 1):
            if norm(ws.cell(r, headers["MODELO"]).value) == norm(name) and norm(ws.cell(r, headers["STATUS"]).value) == "PUBLICADO":
                ws.cell(r, headers["STATUS"]).value = "Arquivado"
        published_at = datetime.now().isoformat(timespec="seconds")
        for component, weight in weights.items():
            row = ws.max_row + 1
            values = {
                "MODELO": name, "VERSAO": version, "STATUS": "Publicado",
                "SUBBUCKET": component, "PESO": float(weight), "PUBLICADO_EM": published_at,
                "PUBLICADO_POR": author.strip() or "Não informado", "MOTIVO": reason.strip(),
                "OBSERVACAO": "Publicado pelo app",
            }
            for col, value in values.items():
                ws.cell(row, headers[col]).value = value
        tmp = path.with_name(path.stem + "__tmp_model" + path.suffix)
        wb.save(tmp)
        tmp.replace(path)
        return True, "Cadastro Mestre sincronizado."
    except PermissionError:
        return False, "Modelo publicado no app, mas o Cadastro Mestre estava aberto no Excel e não pôde ser sincronizado."
    except Exception as exc:
        return False, f"Modelo publicado no app, mas falhou a sincronização com o Cadastro Mestre: {exc}"

def load_published_models(base: dict[str, dict[str, float]], master_path: Path | None = None) -> dict[str, dict[str, float]]:
    merged = deepcopy(base)
    mp = master_path or master_products_path()
    for name, weights in load_models_from_master_sheet(str(mp), *file_cache_signature(mp)).items():
        if weights:
            merged[str(name)] = {str(k): float(v) for k, v in weights.items()}
    if not PUBLISHED_MODELS_PATH.exists():
        return merged
    try:
        payload = json.loads(PUBLISHED_MODELS_PATH.read_text(encoding="utf-8"))
        for name, info in payload.get("models", {}).items():
            weights = info.get("weights", {}) if isinstance(info, dict) else {}
            if weights:
                merged[str(name)] = {str(k): float(v) for k, v in weights.items()}
    except Exception:
        pass
    return merged


def publish_model(name: str, weights: dict[str, float], author: str, reason: str) -> tuple[bool, str]:
    total = sum(float(v or 0) for k, v in weights.items() if norm(k) not in PARENT_WEIGHT_KEYS)
    if abs(total - 1.0) > 0.0005:
        return False, f"A soma dos componentes precisa ser 100,00%. Soma atual: {fmt_pct(total)}"
    DATA_DIR.mkdir(exist_ok=True)
    payload = {"models": {}}
    if PUBLISHED_MODELS_PATH.exists():
        try:
            payload = json.loads(PUBLISHED_MODELS_PATH.read_text(encoding="utf-8"))
        except Exception:
            payload = {"models": {}}
    current = payload.setdefault("models", {}).get(name, {})
    version = int(current.get("version", 0)) + 1
    record = {
        "version": version,
        "published_at": datetime.now().isoformat(timespec="seconds"),
        "published_by": author.strip() or "Não informado",
        "reason": reason.strip(),
        "weights": {str(k): float(v) for k, v in weights.items()},
    }
    payload["models"][name] = record
    PUBLISHED_MODELS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    with MODEL_HISTORY_PATH.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps({"model": name, **record}, ensure_ascii=False) + "\n")
    synced, sync_msg = sync_published_model_to_master(name, weights, version, author, reason)
    return True, f"Modelo publicado na versão {version}. {sync_msg}"


def model_history() -> pd.DataFrame:
    if not MODEL_HISTORY_PATH.exists():
        return pd.DataFrame()
    rows = []
    try:
        for line in MODEL_HISTORY_PATH.read_text(encoding="utf-8").splitlines():
            if line.strip():
                rec = json.loads(line)
                rows.append({k: v for k, v in rec.items() if k != "weights"})
    except Exception:
        return pd.DataFrame()
    return pd.DataFrame(rows)


def theoretical_hierarchy_table(p: dict[str, float], valor: float) -> pd.DataFrame:
    sub = subbucket_targets_from_model(p, valor).rename(columns={"Peso Ideal": "Peso sugerido", "Valor Ideal": "Valor sugerido"})
    rows = []
    groups = [
        ("RENDA FIXA NO BRASIL", ["RF Brasil"]),
        ("RENDA VARIÁVEL NO BRASIL", ["RV Brasil"]),
        ("INVESTIMENTOS INTERNACIONAIS", ["Internacional"]),
        ("ALTERNATIVOS", ["Alternativos"]),
    ]
    for label, classes in groups:
        part = sub[sub["Classe"].isin(classes)]
        if part.empty:
            continue
        rows.append({"Estratégia": label, "Peso sugerido": float(part["Peso sugerido"].sum()), "Valor sugerido": float(part["Valor sugerido"].sum()), "_header": True})
        for _, r in part.iterrows():
            rows.append({"Estratégia": "  " + friendly_strategy_name(r["Subbucket"]), "Peso sugerido": float(r["Peso sugerido"]), "Valor sugerido": float(r["Valor sugerido"]), "_header": False})
    rows.append({"Estratégia": "TOTAL", "Peso sugerido": sum(x["Peso sugerido"] for x in rows if not x["_header"]), "Valor sugerido": sum(x["Valor sugerido"] for x in rows if not x["_header"]), "_header": True})
    return pd.DataFrame(rows)


def theoretical_hierarchy_styler(df: pd.DataFrame):
    view = df[["Estratégia", "Peso sugerido", "Valor sugerido"]].copy()
    header_mask = df["_header"].astype(bool).tolist()
    def row_style(row):
        if header_mask[row.name]:
            return ["background-color: rgba(93,115,170,.34); font-weight:900; border-top:1px solid rgba(255,255,255,.22);" for _ in row]
        return ["" for _ in row]
    return view.style.format({"Peso sugerido": fmt_pct, "Valor sugerido": format_brl}).apply(row_style, axis=1)


# =============================================================================
# Layout global
# =============================================================================
pesos_path = find_file("Pesos-alocacao.xlsx")
pesos_base = load_pesos_xlsx(str(pesos_path), *file_cache_signature(pesos_path))
pesos = load_published_models(pesos_base, master_products_path())
df_contas = load_contas_cached()

lp = logo_path()
if lp:
    h_logo, h_title = st.columns([1.15, 5.85], vertical_alignment="center")
    with h_logo:
        st.image(str(lp), width=250)
    with h_title:
        st.markdown('<h1 class="mw-title">Balanceamento de Carteiras</h1>', unsafe_allow_html=True)
        st.markdown(f'<div class="mw-version">M Wealth • Versão {APP_VERSION}</div>', unsafe_allow_html=True)
else:
    st.markdown('<h1 class="mw-title">M Wealth - Balanceamento de Carteiras</h1>', unsafe_allow_html=True)
    st.markdown(f'<div class="mw-version">Versão {APP_VERSION}</div>', unsafe_allow_html=True)

page = st.segmented_control(
    "",
    ["Controle de Saldo", "Asset Allocation", "Carteira Teórica", "Gestão"],
    default="Controle de Saldo",
)
if page is None:
    page = "Controle de Saldo"

st.markdown('<div class="mw-line"></div>', unsafe_allow_html=True)

# =============================================================================
# Página 1 - Controle de saldo
# =============================================================================
if page == "Controle de Saldo":
    page_intro("Rotina operacional", "Controle de saldo", "Identifique rapidamente contas com caixa disponível, saldos negativos e prioridades de aplicação.", ["Atualização controlada", "Filtro por saldo", "Visão por conta"])

    c0, c1, c2 = st.columns([1.0, 1.35, 4.2], vertical_alignment="bottom")
    force = c0.button("Atualizar base", type="primary", use_container_width=True)
    if force:
        st.cache_data.clear()
    min_saldo_txt = c1.text_input("Saldo mínimo", value=format_brl(10000.0), help="Digite no formato financeiro. Ex.: R$ 10.000,00")
    min_saldo = parse_brl_input(min_saldo_txt, 10000.0)

    try:
        df_latest, meta, mode = load_positions_cached(force_rebuild=force)
        df_latest = enrich_positions_cached(df_latest, mapping_files_signature())
        df_latest = ensure_saldo_operacional(df_latest)
    except Exception as e:
        st.error(f"Falha ao carregar/consolidar posições: {e}")
        st.stop()

    st.caption(f"Última consolidação: **{meta.get('built_at', 'n/d')}**")

    saldo_conta = (
        df_latest[df_latest["saldo_operacional"]]
        .groupby(["GRUPO GERAL", "CLIENTE", "corretora", "conta"], dropna=False, as_index=False)["valor_mercado"]
        .sum()
        .rename(columns={"valor_mercado": "Saldo"})
    )
    pl_conta = (
        df_latest.groupby(["GRUPO GERAL", "CLIENTE", "corretora", "conta"], dropna=False, as_index=False)["valor_mercado"]
        .sum()
        .rename(columns={"valor_mercado": "PL"})
    )
    painel = pl_conta.merge(saldo_conta, how="left", on=["GRUPO GERAL", "CLIENTE", "corretora", "conta"]).fillna({"Saldo": 0.0})

    operaveis = painel[(painel["Saldo"] >= min_saldo) | (painel["Saldo"] < 0)].sort_values("Saldo", ascending=False)

    k1, k2, k3 = st.columns(3)
    with k1:
        metric_card("Contas com saldo acima do mínimo", f"{int((painel['Saldo'] >= min_saldo).sum())}")
    with k2:
        metric_card("Saldo total disponível", format_brl(painel.loc[painel["Saldo"] > 0, "Saldo"].sum()))
    with k3:
        metric_card("Caixa negativo", format_brl(painel.loc[painel["Saldo"] < 0, "Saldo"].sum()))

    st.subheader("Contas com saldo para operação")
    cols = ["GRUPO GERAL", "CLIENTE", "corretora", "conta", "PL", "Saldo"]
    st.dataframe(prepare_display(operaveis[cols], money_cols=["PL", "Saldo"], max_rows=700), use_container_width=True, hide_index=True)

    with st.expander("Ver todas as contas, inclusive sem saldo operacional", expanded=False):
        st.dataframe(prepare_display(painel[cols].sort_values("Saldo", ascending=False), money_cols=["PL", "Saldo"], max_rows=1200), use_container_width=True, hide_index=True)


# =============================================================================
# Página 2 - Asset Allocation
# =============================================================================
if page == "Asset Allocation":
    page_intro("Diagnóstico e execução", "Asset Allocation", "Compare a carteira atual com o modelo, entenda os desvios e transforme necessidades por estratégia em recomendações executáveis.", ["Atual x ideal", "Produtos exatos", "Restrições", "Qualidade da base"])
    cadastro_ativo = master_products_path()
    st.caption(f"Cadastro mestre carregado: **{cadastro_ativo.name}**")
    cadastro_sig = file_cache_signature(cadastro_ativo)
    pool_produtos = load_product_pool_cached(str(cadastro_ativo), *cadastro_sig)
    restricoes_base = load_client_restrictions_cached(str(cadastro_ativo), *cadastro_sig)
    try:
        _b3_health = load_b3_master_cached(str(cadastro_ativo), *file_cache_signature(cadastro_ativo))
        _fund_health = load_fundos_prev_cached(str(cadastro_ativo), *file_cache_signature(cadastro_ativo))
        with st.expander("Saúde das bases e cadastros", expanded=False):
            h1, h2, h3, h4 = st.columns(4)
            h1.metric("Ativos B3 cadastrados", len(_b3_health))
            h2.metric("Fundos/Previdência", len(_fund_health))
            h3.metric("Arquivo mestre", cadastro_ativo.name)
            h4.metric("Pesos carregados", len(pesos))
            invalid_models = []
            for _model_name, _model_weights in pesos.items():
                _total_children = sum(float(v or 0) for k, v in _model_weights.items() if norm(k) not in PARENT_WEIGHT_KEYS)
                if abs(_total_children - 1.0) > 0.015:
                    invalid_models.append(f"{_model_name}: {_total_children:.2%}")
            if invalid_models:
                st.warning("Modelos cuja soma dos componentes difere de 100%: " + ", ".join(invalid_models))
            if _b3_health.empty:
                st.error("A aba Ativos B3 não foi carregada ou não possui registros válidos.")
            if _fund_health.empty:
                st.warning("As abas de Fundos/Previdência estão vazias ou não puderam ser lidas.")
    except Exception as _health_error:
        st.error(f"Falha ao validar o cadastro mestre: {_health_error}")
    try:
        df_latest, meta, mode = load_positions_cached(force_rebuild=False)
        df_latest = enrich_positions_cached(df_latest, mapping_files_signature())
        df_latest = ensure_saldo_operacional(df_latest)
    except Exception as e:
        st.error(f"Não consegui carregar a base: {e}")
        st.stop()

    if "GRUPO GERAL" not in df_latest.columns:
        st.error("A base não possui a coluna GRUPO GERAL. Verifique o arquivo Contas.xlsx.")
        st.stop()
    if not pesos:
        st.error("Pesos-alocacao.xlsx não foi encontrado ou não pôde ser lido.")
        st.stop()

    grupos = sorted(df_latest["GRUPO GERAL"].dropna().astype(str).unique())
    col_g, col_c, col_m = st.columns([3, 3, 2])
    with col_g:
        grupo_sel = st.selectbox("Grupo familiar", grupos)
    contas_info = df_latest[df_latest["GRUPO GERAL"].astype(str).eq(str(grupo_sel))][["conta", "CLIENTE", "corretora"]].drop_duplicates()
    with col_c:
        opcoes = ["Todas as contas"] + [f"{str(r.CLIENTE).strip()} • {r.corretora} ({r.conta})" if pd.notna(r.CLIENTE) else f"{r.corretora} ({r.conta})" for r in contas_info.itertuples(index=False)]
        conta_sel = st.selectbox("Conta", opcoes)
    with col_m:
        perfil_cliente = "Não identificado"
        if not df_contas.empty and "GRUPO GERAL" in df_contas.columns:
            m = df_contas[df_contas["GRUPO GERAL"].astype(str).str.strip().eq(str(grupo_sel).strip())]
            if not m.empty and "Perfil Carteira" in m.columns:
                perfil_cliente = str(m["Perfil Carteira"].iloc[0]).strip()
        modelos = list(pesos.keys())
        idx = modelos.index(model_for_profile(perfil_cliente, modelos)) if model_for_profile(perfil_cliente, modelos) in modelos else 0
        modelo = st.selectbox("Modelo de alocação", modelos, index=idx)

    pos_cliente = df_latest[df_latest["GRUPO GERAL"].astype(str).eq(str(grupo_sel))].copy()
    conta_real = ""
    cliente_referencia = ""
    if conta_sel != "Todas as contas":
        conta_real = conta_sel.split("(")[-1].strip(")")
        pos_cliente = pos_cliente[pos_cliente["conta"].astype(str).eq(conta_real)].copy()
    if not pos_cliente.empty and "CLIENTE" in pos_cliente.columns:
        cliente_referencia = str(pos_cliente["CLIENTE"].dropna().astype(str).iloc[0]) if pos_cliente["CLIENTE"].notna().any() else ""
    restricoes_cliente = applicable_restrictions(restricoes_base, str(grupo_sel), cliente_referencia, conta_real)

    p = pesos[modelo]

    base_mode = st.radio(
        "Base usada no balanceamento",
        ["PL patrimonial total", "PL investível"],
        horizontal=True,
        help="PL investível exclui caixa, proventos e ativos fora da estratégia do denominador dos alvos.",
    )

    # Marca somente as posições reais de bolsa pelo preço atual do Yahoo Finance.
    # O PL, os valores atuais e as diferenças passam a refletir quantidade × cotação atual.
    held_tickers = pos_cliente.loc[exchange_position_mask(pos_cliente), "ticker_norm"].dropna().astype(str).tolist()
    model_tickers = [t for xs in rv_universe(modelo).values() for t in xs] + FI_INFRA_TICKERS
    quote_tickers = tuple(sorted({ticker_clean(t) for t in held_tickers + model_tickers if ticker_clean(t)}))
    price_ref = load_yfinance_prices(quote_tickers)
    pos_cliente = mark_exchange_positions_to_market(pos_cliente, price_ref)

    pl_patrimonial = float(pos_cliente["valor_mercado"].sum())
    investible_mask = ~pos_cliente["classe_macro"].isin(["Caixa", "Fora da Estratégia"])
    pl_investivel = float(pos_cliente.loc[investible_mask, "valor_mercado"].sum())
    pl = pl_patrimonial if base_mode == "PL patrimonial total" else pl_investivel
    pos_calculo = pos_cliente if base_mode == "PL patrimonial total" else pos_cliente.loc[investible_mask].copy()
    macro_df, sub_df = portfolio_tables(pos_calculo, p, pl)
    corretoras_cliente = set(pos_cliente.get("corretora", pd.Series(dtype=str)).dropna().astype(str).unique())
    recomendacoes_produtos = recommend_exact_products(sub_df, pos_cliente, pool_produtos, restricoes_cliente, pl, corretoras_cliente)

    pl_xp = float(pos_cliente.loc[pos_cliente["corretora"].eq("XP"), "valor_mercado"].sum())
    pl_btg = float(pos_cliente.loc[pos_cliente["corretora"].eq("BTG"), "valor_mercado"].sum())
    pl_cs = float(pos_cliente.loc[pos_cliente["corretora"].eq("CS"), "valor_mercado"].sum())
    saldo = float(pos_cliente.loc[pos_cliente.get("saldo_operacional", False).fillna(False).astype(bool), "valor_mercado"].sum()) if "saldo_operacional" in pos_cliente.columns else float(pos_cliente.loc[pos_cliente["subbucket"].eq("Saldo em Conta"), "valor_mercado"].sum())
    manual_matches = int(pos_cliente.get("manual_match", pd.Series([False] * len(pos_cliente), index=pos_cliente.index)).fillna(False).astype(bool).sum())
    nao_class = float(pos_cliente.loc[pos_cliente["subbucket"].eq("Outros / Não Classificado"), "valor_mercado"].sum())

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    with k1:
        metric_card("PL Patrimonial", format_brl(pl_patrimonial))
    with k2:
        metric_card("Base de Alocação", format_brl(pl))
    with k3:
        metric_card("XP", format_brl(pl_xp))
    with k4:
        metric_card("BTG", format_brl(pl_btg))
    with k5:
        metric_card("CS", format_brl(pl_cs))
    with k6:
        metric_card("Saldo", format_brl(saldo))
    st.caption(f"Perfil: **{perfil_cliente}** • Modelo aplicado: **{modelo}** • Fundos reconhecidos pelo manual: **{manual_matches}** • Não classificado: **{format_brl(nao_class)}**")

    st.markdown('<div class="mw-line"></div>', unsafe_allow_html=True)

    st.subheader("Visão macro atual x ideal")
    macro_view = macro_df.copy()
    macro_view["Classe"] = macro_view["Classe"].apply(friendly_class_name)
    macro_view = macro_view.drop(columns=["Ação", "Status"], errors="ignore")

    macro_hier = macro_hierarchy_table(sub_df, pl)
    col1, col2 = st.columns([0.95, 2.45])
    with col1:
        # Mesmo conceito da carteira teórica: gráfico pela abertura das estratégias,
        # não apenas pelo macro Renda Fixa / Renda Variável / Internacional.
        plot = macro_hier.copy()
        plot = plot[(~plot.get("_header", False).astype(bool)) & (plot["Quanto tem"] > 0)].copy()
        plot["Estratégia"] = plot["Estratégia"].astype(str).str.strip()
        if plot.empty:
            plot = macro_df[macro_df["Valor Atual"] > 0].copy()
            plot["Estratégia"] = plot["Classe"].apply(friendly_class_name)
            plot = plot.rename(columns={"Valor Atual": "Quanto tem"})
        fig = px.pie(plot, names="Estratégia", values="Quanto tem", title="Distribuição atual por estratégia", hole=.48)
        fig.update_layout(height=365, margin=dict(l=8, r=8, t=45, b=8), showlegend=True, legend_title_text="")
        st.plotly_chart(fig, use_container_width=True)
    with col2:
        st.dataframe(
            macro_hierarchy_styler(macro_hier),
            use_container_width=True,
            hide_index=True,
            height=min(820, max(420, 37 * (len(macro_hier) + 1))),
        )

    st.subheader("Abertura por estratégia")
    st.markdown('<div class="mw-muted">Abertura objetiva por classe. A diferença positiva indica valor a alocar; diferença negativa indica excesso.</div>', unsafe_allow_html=True)
    class_order = ["RF Brasil", "Alternativos"]
    for classe in class_order:
        class_df = sub_df[sub_df["Classe"].eq(classe)].copy()
        if class_df.empty:
            continue
        valor_atual_cls = float(class_df["Valor Atual"].sum())
        valor_ideal_cls = float(class_df["Valor Ideal"].sum())
        diff_cls = float(class_df["Diferença"].sum())
        titulo = f"{friendly_class_name(classe)} • Atual {format_brl_label(valor_atual_cls)} | Ideal {format_brl_label(valor_ideal_cls)} | Diferença {format_brl_label(diff_cls)}"
        expanded = classe == "RF Brasil" or abs(diff_cls) > 300
        with st.expander(titulo, expanded=expanded):
            class_view = class_df.copy()
            class_view["Estratégia"] = class_view["Subbucket"].apply(friendly_strategy_name)
            class_view = class_view[["Estratégia", "Peso Atual", "Peso Ideal", "Valor Atual", "Valor Ideal", "Diferença"]]
            st.dataframe(
                money_color_styler(
                    class_view,
                    money_cols=["Valor Atual", "Valor Ideal", "Diferença"],
                    pct_cols=["Peso Atual", "Peso Ideal"],
                    diff_cols=["Diferença"],
                ),
                use_container_width=True,
                hide_index=True,
            )

            ativos_cls = pos_cliente[pos_cliente["classe_macro"].eq(classe)].copy()
            if not ativos_cls.empty:
                cols_pos = ["subbucket", "asset_id", "asset_nome", "corretora", "valor_mercado", "quantidade", "manual_fundo", "manual_liquidez", "b3_estrategia", "b3_liquidez_operacional", "fonte_preco", "rebalancear", "tratamento"]
                ativos_cls = ativos_cls[[c for c in cols_pos if c in ativos_cls.columns]].sort_values("valor_mercado", ascending=False).head(80)
                ativos_cls["subbucket"] = ativos_cls["subbucket"].apply(friendly_strategy_name)
                ativos_cls = ativos_cls.rename(columns={
                    "subbucket": "Estratégia",
                    "asset_id": "Ativo",
                    "asset_nome": "Nome",
                    "corretora": "Corretora",
                    "valor_mercado": "Valor",
                    "quantidade": "Quantidade",
                    "manual_fundo": "Fundo no manual",
                    "manual_liquidez": "Liquidez D+",
                    "b3_estrategia": "Estratégia cadastrada",
                    "b3_liquidez_operacional": "Liquidez operacional B3",
                    "fonte_preco": "Fonte de preço",
                    "rebalancear": "Rebalancear",
                    "tratamento": "Origem do match",
                })
                with st.expander("Ver ativos classificados nessa classe", expanded=False):
                    st.dataframe(
                        prepare_display(ativos_cls, money_cols=["Valor"], qty_cols=["Quantidade"], max_rows=80),
                        use_container_width=True,
                        hide_index=True,
                    )

    section_title("Produtos recomendados para execução", "O motor cruza a necessidade por estratégia com o pool elegível, corretora, mínimos, limites e restrições do cliente.")
    if recomendacoes_produtos.empty:
        st.info("Nenhum produto exato foi sugerido. Preencha o Pool de Produtos para os subbuckets que precisam de compra.")
    else:
        st.dataframe(
            money_color_styler(
                recomendacoes_produtos,
                money_cols=["Valor atual no produto", "Alvo do produto", "Valor recomendado"],
                pct_cols=["Peso no pool", "Limite"],
                diff_cols=["Valor recomendado"],
            ),
            use_container_width=True,
            hide_index=True,
        )
        st.download_button(
            "Baixar recomendação de produtos",
            data=dataframe_excel_bytes(recomendacoes_produtos, "Produtos recomendados"),
            file_name=f"produtos_recomendados_{str(grupo_sel).replace(' ', '_').lower()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    section_title("Produtos de bolsa e infraestrutura", "Ativos listados utilizam preço de mercado quando permitido pelo cadastro mestre.")
    if not HAS_YFINANCE:
        st.error("A biblioteca yfinance não está instalada. Instale-a para habilitar preços atuais e cálculos de ordens.")
    missing_quotes = sorted(set(quote_tickers) - set(price_ref))
    if missing_quotes:
        st.warning(
            "Sem cotação do Yahoo Finance para: " + ", ".join(missing_quotes)
            + ". Esses ativos ficarão sem recomendação de quantidade até a cotação estar disponível."
        )
    else:
        st.caption("Preços atualizados pelo Yahoo Finance, com cache de 5 minutos. Valor atual = quantidade real × preço de mercado.")
    rv_df = apply_restrictions_to_market_orders(
        rv_recommendation(pos_cliente, p, pl, modelo, price_ref), restricoes_cliente, pool_produtos, pl
    )
    fi_df = apply_restrictions_to_market_orders(
        fiinfra_recommendation(pos_cliente, p, pl, price_ref), restricoes_cliente, pool_produtos, pl
    )
    tab_a, tab_b = st.tabs(["Ações e Fundos Imobiliários / FIAGROs", "Infraestrutura"])
    with tab_a:
        rv_view = rv_df[rv_df["Valor Ideal"].gt(0) | rv_df["Valor Atual"].gt(0)].copy()
        rv_table = rv_view.drop(columns=["Grupo"], errors="ignore")
        if rv_view.empty:
            st.info("Este modelo não possui alvo para ações ou fundos imobiliários.")
        else:
            st.dataframe(
                money_color_styler(
                    rv_table,
                    money_cols=["Preço referência", "Valor Atual", "Valor Ideal", "Diferença"],
                    qty_cols=["Qtd Atual", "Qtd Ideal", "Qtd a operar"],
                    diff_cols=["Diferença", "Qtd a operar"],
                ),
                use_container_width=True,
                hide_index=True,
            )
            basket = basket_excel_bytes(rv_view, grupo_sel)
            st.download_button(
                "Baixar basket de ações e fundos imobiliários",
                data=basket,
                file_name=f"basket_acoes_fiis_{str(grupo_sel).replace(' ', '_').lower()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
    with tab_b:
        fi_view = fi_df[fi_df["Valor Ideal"].gt(0) | fi_df["Valor Atual"].gt(0)].copy()
        fi_table = fi_view.drop(columns=["Grupo"], errors="ignore")
        if fi_view.empty:
            st.info("Este modelo não possui alvo para infraestrutura.")
        else:
            st.dataframe(
                money_color_styler(
                    fi_table,
                    money_cols=["Preço referência", "Valor Atual", "Valor Ideal", "Diferença"],
                    qty_cols=["Qtd Atual", "Qtd Ideal", "Qtd a operar"],
                    diff_cols=["Diferença", "Qtd a operar"],
                ),
                use_container_width=True,
                hide_index=True,
            )
            basket = basket_excel_bytes(fi_view, grupo_sel)
            st.download_button(
                "Baixar basket de infraestrutura",
                data=basket,
                file_name=f"basket_infra_{str(grupo_sel).replace(' ', '_').lower()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.subheader("Detalhamento das posições")
    buckets = [b for b in SUBBUCKET_ORDER if b in set(pos_cliente["subbucket"])]
    bucket_options = {friendly_strategy_name(b): b for b in buckets}
    if bucket_options:
        bucket_label = st.selectbox("Abrir posições por estratégia", list(bucket_options.keys()))
        bucket_sel = bucket_options[bucket_label]
        subpos = pos_cliente[pos_cliente["subbucket"].eq(bucket_sel)].copy()
        agrup = subpos.groupby(["asset_id", "asset_nome", "corretora", "subbucket"], dropna=False, as_index=False).agg(Valor=("valor_mercado", "sum"), Quantidade=("quantidade", "sum"), Contas=("conta", "nunique")).sort_values("Valor", ascending=False)
        agrup["Peso no Cliente"] = np.where(pl > 0, agrup["Valor"] / pl, 0)
        agrup["Estratégia"] = agrup["subbucket"].apply(friendly_strategy_name)
        agrup = agrup.rename(columns={"asset_id": "Ativo", "asset_nome": "Nome", "corretora": "Corretora"})
        st.dataframe(
            prepare_display(agrup[["Estratégia", "Ativo", "Nome", "Corretora", "Valor", "Quantidade", "Peso no Cliente", "Contas"]], money_cols=["Valor"], pct_cols=["Peso no Cliente"], qty_cols=["Quantidade"], max_rows=600),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.info("Nenhuma estratégia encontrada para detalhamento.")

    with st.expander("Revisões e exceções", expanded=False):
        universo = set(ACOES_SEM_RENDA + ACOES_COM_RENDA + FIIS_RECOMENDADOS + FI_INFRA_TICKERS + list(ATIVOS_ESTRATEGICOS_B3.keys()))
        fora_df = pos_cliente[
            pos_cliente["classe_macro"].eq("Fora da Estratégia") |
            ((pos_cliente["classe_macro"].eq("RV Brasil")) & (~pos_cliente["ticker_norm"].isin({ticker_clean(x) for x in universo}))) |
            (pos_cliente["subbucket"].str.contains("Sem Liquidez|Não Classificado|COE|Previdência", case=False, na=False))
        ].sort_values("valor_mercado", ascending=False)
        cols = ["corretora", "conta", "CLIENTE", "asset_id", "asset_nome", "classe_macro", "subbucket", "tratamento", "manual_fundo", "manual_classe", "manual_liquidez", "manual_metodo", "manual_score", "b3_match", "b3_tipo_produto", "b3_estrategia", "b3_liquidez_operacional", "b3_fonte_preco", "fonte_preco", "liquidez_operacional_aplicada", "b3_status_mapeamento", "rebalancear", "valor_mercado", "quantidade", "indexador", "liquidez", "vencimento"]
        view = fora_df[[c for c in cols if c in fora_df.columns]].copy()
        for c in ["classe_macro", "subbucket"]:
            if c in view.columns:
                view[c] = view[c].apply(friendly_class_name if c == "classe_macro" else friendly_strategy_name)
        st.dataframe(prepare_display(view, money_cols=["valor_mercado"], qty_cols=["quantidade"], max_rows=500), use_container_width=True, hide_index=True)


# =============================================================================
# Página 3 - Carteira Teórica
# =============================================================================
if page == "Carteira Teórica":
    page_intro("Simulação para cliente", "Carteira teórica", "Visualize a carteira modelo com a mesma lógica do Asset Allocation e gere um material institucional pronto para apresentação.", ["Tabela hierárquica", "Produtos da estratégia", "PDF institucional"])
    modelos = list(pesos.keys())
    if not modelos:
        st.error("Pesos-alocacao.xlsx não foi encontrado ou não pôde ser lido.")
        st.stop()

    c1, c2, c3 = st.columns([2, 1.2, 2])
    with c1:
        modelo = st.selectbox("Perfil de carteira", modelos)
    with c2:
        valor = st.number_input("Valor simulado", min_value=0.0, value=1_000_000.0, step=100_000.0, format="%.2f")
    with c3:
        cliente = st.text_input("Nome do cliente no PDF (opcional)")

    df_teor = theoretical_portfolio(pesos[modelo], valor, modelo)
    if df_teor.empty:
        st.warning("Não encontrei componentes válidos para essa carteira. Verifique a planilha de pesos.")
        st.stop()

    macro = portfolio_macro_cliente(df_teor)
    renda_fixa = macro[macro["Classe de investimento"].str.contains("Renda fixa|infraestrutura|Crédito privado", case=False, na=False)]["Valor sugerido"].sum()
    renda_variavel = macro[macro["Classe de investimento"].str.contains("Ações|Fundos Imobiliários|FIAGROs", case=False, na=False)]["Valor sugerido"].sum()
    internacional = macro[macro["Classe de investimento"].str.contains("internacionais", case=False, na=False)]["Valor sugerido"].sum()

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Valor simulado", format_brl(valor))
    k2.metric("Renda fixa no Brasil", format_brl(renda_fixa))
    k3.metric("Renda variável no Brasil", format_brl(renda_variavel))
    k4.metric("Investimentos internacionais", format_brl(internacional))

    col_a, col_b = st.columns([1.05, 1.2])
    with col_a:
        fig = px.pie(macro, names="Classe de investimento", values="Valor sugerido", title="Distribuição sugerida", hole=.48)
        fig.update_layout(height=330, margin=dict(l=8, r=8, t=45, b=8), legend_title_text="")
        st.plotly_chart(fig, use_container_width=True)
    with col_b:
        section_title("Resumo da alocação", "Mesma hierarquia utilizada no Asset Allocation.")
        teor_hier = theoretical_hierarchy_table(pesos[modelo], valor)
        st.dataframe(
            theoretical_hierarchy_styler(teor_hier),
            use_container_width=True,
            hide_index=True,
            height=min(760, max(390, 36 * (len(teor_hier) + 1))),
        )

    st.markdown('<div class="mw-line"></div>', unsafe_allow_html=True)
    st.subheader("Composição por classe de investimento")

    for group in macro["Classe de investimento"].tolist():
        comp = df_teor[(df_teor["Grupo"].eq(group)) & (df_teor["Nível"].eq("Composição"))].copy()
        ativos = df_teor[(df_teor["Grupo"].eq(group)) & (df_teor["Nível"].eq("Ativo"))].copy()
        total_group = float(comp["Valor"].sum())
        peso_group = float(comp["Peso"].sum())
        with st.expander(f"{group} — {fmt_pct(peso_group)} | {format_brl(total_group)}", expanded=True):
            st.markdown(f'<div class="mw-muted">{GROUP_DESCRIPTIONS.get(group, "Componente da estratégia de alocação.")}</div>', unsafe_allow_html=True)
            tabela_comp = comp.rename(columns={"Peso": "Peso sugerido", "Valor": "Valor sugerido"})[["Composição", "Peso sugerido", "Valor sugerido", "Explicação"]]
            st.dataframe(
                prepare_display(tabela_comp, money_cols=["Valor sugerido"], pct_cols=["Peso sugerido"]),
                use_container_width=True,
                hide_index=True,
            )
            if not ativos.empty:
                st.markdown("**Ativos utilizados na carteira modelo**")
                if ativos["Composição"].nunique() > 1:
                    for comp_nome, ativos_comp in ativos.groupby("Composição", sort=False):
                        st.markdown(f"_{comp_nome}_")
                        tabela_ativos = ativos_comp.rename(columns={"Peso": "Peso sugerido", "Valor": "Valor sugerido"})[["Ativo", "Peso sugerido", "Valor sugerido"]]
                        st.dataframe(
                            prepare_display(tabela_ativos, money_cols=["Valor sugerido"], pct_cols=["Peso sugerido"]),
                            use_container_width=True,
                            hide_index=True,
                        )
                else:
                    tabela_ativos = ativos.rename(columns={"Peso": "Peso sugerido", "Valor": "Valor sugerido"})[["Ativo", "Peso sugerido", "Valor sugerido"]]
                    st.dataframe(
                        prepare_display(tabela_ativos, money_cols=["Valor sugerido"], pct_cols=["Peso sugerido"]),
                        use_container_width=True,
                        hide_index=True,
                    )

    st.markdown('<div class="mw-line"></div>', unsafe_allow_html=True)
    st.subheader("Gerar material para cliente")
    try:
        pdf = build_pdf_teorico(df_teor, modelo, valor, cliente)
        st.download_button(
            "Baixar PDF da carteira teórica",
            data=pdf,
            file_name=f"carteira_teorica_mwealth_{modelo.lower().replace(' ', '_')}.pdf",
            mime="application/pdf",
            type="primary",
        )
    except Exception as e:
        st.info(f"PDF indisponível: {e}")


# =============================================================================
# Página 4 - Gestão
# =============================================================================
if page == "Gestão":
    page_intro("Governança", "Gestão de modelos e personalizações", "Edite pesos, pools, restrições e exceções operacionais diretamente no app. As alterações são registradas no Cadastro Mestre e passam a valer após o salvamento/publicação.", ["Modelos versionados", "Input direto", "Restrições", "Cadastro Mestre"])
    cadastro_ativo = master_products_path()
    sig = file_cache_signature(cadastro_ativo)
    pool_admin = load_product_pool_cached(str(cadastro_ativo), *sig)
    restrictions_admin = load_client_restrictions_cached(str(cadastro_ativo), *sig)
    b3_admin = load_b3_master_cached(str(cadastro_ativo), *sig)
    fundos_admin = load_fundos_prev_cached(str(cadastro_ativo), *sig)

    k1, k2, k3, k4 = st.columns(4)
    with k1: metric_card("Produtos no pool", str(len(pool_admin)))
    with k2: metric_card("Restrições ativas/cadastradas", str(len(restrictions_admin)))
    with k3: metric_card("Ativos B3", str(len(b3_admin)))
    with k4: metric_card("Fundos e previdência", str(len(fundos_admin)))


    tab_modelos, tab_personalizacoes, tab_historico = st.tabs(["Modelos de alocação", "Personalizações e regras", "Histórico de publicação"])
    with tab_modelos:
        section_title("Editar e publicar modelo", "A publicação grava a versão no app e também sincroniza a aba Modelos de Alocação do Cadastro Mestre.")
        nomes_modelos = list(pesos.keys())
        modelo_admin = st.selectbox("Modelo", nomes_modelos, key="modelo_admin")
        base_weights = pesos[modelo_admin]
        editable = pd.DataFrame([{"Componente": k, "Peso": float(v)} for k, v in base_weights.items() if norm(k) not in PARENT_WEIGHT_KEYS])
        edited = st.data_editor(
            editable, use_container_width=True, hide_index=True, disabled=["Componente"],
            column_config={"Peso": st.column_config.NumberColumn("Peso", min_value=0.0, max_value=1.0, step=0.005, format="%.4f")},
            key=f"editor_{modelo_admin}",
        )
        total = float(pd.to_numeric(edited["Peso"], errors="coerce").fillna(0).sum())
        c1, c2, c3 = st.columns([1, 1, 2])
        with c1: metric_card("Soma do modelo", fmt_pct(total))
        with c2: metric_card("Status", "Válido" if abs(total - 1) <= .0005 else "Revisar")
        with c3:
            if abs(total - 1) > .0005: st.warning("A soma precisa ser exatamente 100% para publicar.")
            else: st.success("Modelo validado e pronto para publicação.")
        author = st.text_input("Responsável pela publicação")
        reason = st.text_area("Motivo da alteração", placeholder="Ex.: atualização do comitê de alocação de agosto/2026")
        if st.button("Publicar nova versão", type="primary", use_container_width=True, disabled=abs(total - 1) > .0005):
            new_weights = {str(r["Componente"]): float(r["Peso"]) for _, r in edited.iterrows()}
            ok, message = publish_model(modelo_admin, new_weights, author, reason)
            if ok:
                st.success(message + " A nova versão já será usada após a recarga.")
                st.cache_data.clear()
            else: st.error(message)

    with tab_personalizacoes:
        st.info("As edições abaixo alteram o Cadastro Mestre. Para evitar conflito de arquivo, mantenha a planilha fechada no Excel enquanto salvar pelo app.")
        p_pool, p_restr, p_b3, p_fundos = st.tabs(["Pool de Produtos", "Restrições por Cliente", "Ativos B3", "Overrides de Fundos/Prev"])

        with p_pool:
            section_title("Pool de produtos", "Defina exatamente quais produtos podem receber cada necessidade de alocação.")
            pool_raw = read_master_sheet_for_editor("Pool de Produtos", POOL_COLUMNS)
            pool_edit = st.data_editor(
                pool_raw, num_rows="dynamic", use_container_width=True, hide_index=True,
                column_config={
                    "PESO_NO_POOL": st.column_config.NumberColumn("PESO_NO_POOL", min_value=0.0, max_value=1.0, step=0.05, format="%.2f"),
                    "LIMITE_POR_CLIENTE": st.column_config.NumberColumn("LIMITE_POR_CLIENTE", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
                    "APORTE_MINIMO": st.column_config.NumberColumn("APORTE_MINIMO", min_value=0.0, step=100.0),
                    "PRIORIDADE_COMPRA": st.column_config.NumberColumn("PRIORIDADE_COMPRA", min_value=1, step=1),
                }, key="pool_editor_direct",
            )
            if st.button("Salvar Pool de Produtos", type="primary", use_container_width=True):
                ok, msg = save_master_sheet_from_app("Pool de Produtos", pool_edit, POOL_COLUMNS, "NOME_PRODUTO")
                (st.success if ok else st.error)(msg)

        with p_restr:
            section_title("Restrições por cliente", "Cadastre regras por grupo, cliente ou conta. Classe, subbucket, setor, gestora, ticker, CNPJ e produto são suportados.")
            restr_raw = read_master_sheet_for_editor("Restrições por Cliente", RESTRICTION_COLUMNS)
            restr_edit = st.data_editor(
                restr_raw, num_rows="dynamic", use_container_width=True, hide_index=True,
                column_config={
                    "NIVEL": st.column_config.SelectboxColumn("NIVEL", options=["Grupo", "Cliente", "Conta"]),
                    "TIPO_REGRA": st.column_config.SelectboxColumn("TIPO_REGRA", options=["Ticker", "CNPJ", "Produto", "Setor", "Gestora", "Subbucket", "Classe"]),
                    "ACAO": st.column_config.SelectboxColumn("ACAO", options=["Não comprar", "Não vender", "Manter posição", "Excluir da carteira", "Limitar percentual", "Substituir por produto"]),
                    "LIMITE_PERCENTUAL": st.column_config.NumberColumn("LIMITE_PERCENTUAL", min_value=0.0, max_value=1.0, step=0.01, format="%.2f"),
                    "STATUS": st.column_config.SelectboxColumn("STATUS", options=["Ativa", "Inativa"]),
                }, key="restriction_editor_direct",
            )
            if st.button("Salvar Restrições", type="primary", use_container_width=True):
                ok, msg = save_master_sheet_from_app("Restrições por Cliente", restr_edit, RESTRICTION_COLUMNS, "GRUPO_CLIENTE")
                (st.success if ok else st.error)(msg)

        with p_b3:
            section_title("Ativos B3", "Use para exceções explícitas de ticker, fonte de preço, liquidez e classificação.")
            b3_raw = read_master_sheet_for_editor("Ativos B3", B3_EDITOR_COLUMNS)
            b3_edit = st.data_editor(
                b3_raw, num_rows="dynamic", use_container_width=True, hide_index=True,
                column_config={
                    "REBALANCEAR": st.column_config.SelectboxColumn("REBALANCEAR", options=["Sim", "Não"]),
                    "FONTE_PRECO": st.column_config.SelectboxColumn("FONTE_PRECO", options=["Yahoo Finance", "Valor da posição", "Não precificar"]),
                    "LIQUIDEZ_OPERACIONAL": st.column_config.NumberColumn("LIQUIDEZ_OPERACIONAL", min_value=0.0, step=1.0),
                }, key="b3_editor_direct",
            )
            if st.button("Salvar Ativos B3", type="primary", use_container_width=True):
                ok, msg = save_master_sheet_from_app("Ativos B3", b3_edit, B3_EDITOR_COLUMNS, "TICKER")
                (st.success if ok else st.error)(msg)

        with p_fundos:
            section_title("Override operacional de fundos e previdência", "Selecione um cadastro e altere apenas os campos operacionais. Os dados originais de classificação e liquidez continuam preservados.")
            origem_override = st.radio("Base", ["Fundos de Investimentos", "Previdência"], horizontal=True)
            if origem_override == "Fundos de Investimentos":
                raw_override = pd.read_excel(cadastro_ativo, sheet_name=origem_override)
                name_col = "NOME_FUNDO"
            else:
                raw_override = pd.read_excel(cadastro_ativo, sheet_name=origem_override)
                name_col = "Nome do Fundo Investido pelos planos"
            raw_override.columns = [str(c).strip() for c in raw_override.columns]
            if name_col in raw_override.columns and not raw_override.empty:
                search = st.text_input("Buscar fundo", key="search_override").strip()
                candidates = raw_override.copy()
                if search:
                    candidates = candidates[candidates[name_col].fillna("").astype(str).str.contains(search, case=False, na=False)]
                labels = [f"{idx} • {str(r[name_col])}" for idx, r in candidates.head(200).iterrows()]
                if labels:
                    selected = st.selectbox("Fundo", labels)
                    source_idx = int(selected.split(" • ", 1)[0])
                    row = raw_override.loc[source_idx]
                    c1, c2, c3 = st.columns(3)
                    classe_op = c1.text_input("CLASSE_OPERACIONAL", value=optional_text(row.get("CLASSE_OPERACIONAL", "")), key=f"ov_classe_{source_idx}_{origem_override}")
                    subbucket_op = c2.text_input("SUBBUCKET_OPERACIONAL", value=optional_text(row.get("SUBBUCKET_OPERACIONAL", "")), key=f"ov_sub_{source_idx}_{origem_override}")
                    liq_default = pd.to_numeric(row.get("LIQUIDEZ_OPERACIONAL", np.nan), errors="coerce")
                    liquidez_op = c3.number_input("LIQUIDEZ_OPERACIONAL", min_value=0.0, value=float(liq_default) if pd.notna(liq_default) else 0.0, step=1.0, key=f"ov_liq_{source_idx}_{origem_override}")
                    c4, c5 = st.columns(2)
                    rebalancear_op = c4.selectbox("REBALANCEAR", ["", "Sim", "Não"], index=0, key=f"ov_reb_{source_idx}_{origem_override}")
                    obs_op = c5.text_input("OBSERVACAO_OPERACIONAL", value=optional_text(row.get("OBSERVACAO_OPERACIONAL", "")), key=f"ov_obs_{source_idx}_{origem_override}")
                    if st.button("Salvar override do fundo", type="primary", use_container_width=True):
                        updates = {
                            "CLASSE_OPERACIONAL": classe_op,
                            "SUBBUCKET_OPERACIONAL": subbucket_op,
                            "LIQUIDEZ_OPERACIONAL": liquidez_op if liquidez_op > 0 else None,
                            "REBALANCEAR": rebalancear_op,
                            "STATUS_MAPEAMENTO": "Validado",
                            "OBSERVACAO_OPERACIONAL": obs_op,
                        }
                        ok, msg = update_fund_override_from_app(origem_override, source_idx, updates)
                        (st.success if ok else st.error)(msg)
                else:
                    st.info("Nenhum fundo encontrado para a busca.")
            else:
                st.warning("A aba selecionada não pôde ser lida.")

    with tab_historico:
        section_title("Histórico de versões publicadas")
        hist = model_history()
        if hist.empty: st.info("Nenhuma versão foi publicada pelo app até o momento.")
        else: st.dataframe(hist.sort_values("published_at", ascending=False), use_container_width=True, hide_index=True)


st.caption("M Wealth Asset Allocation")
